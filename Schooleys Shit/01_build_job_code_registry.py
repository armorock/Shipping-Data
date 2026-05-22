"""
Phase 1 — Job Code Registry Builder
Cross-references BOM, Dispatch Board, ERP Shipping, and Job Code Master List
to produce Job_Code_Registry.xlsx and populate Structure Registry sections
in existing Job Code markdown files.
"""

import csv
import os
from collections import defaultdict
from datetime import datetime, date

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
SRC_DIR      = os.path.join(BASE_DIR, "MASTER CSV FILES")
JC_DIR       = os.path.join(BASE_DIR, "Job Codes")
OUTPUT_PATH  = os.path.join(BASE_DIR, "Job_Code_Registry.xlsx")

BOM_PATH      = os.path.join(SRC_DIR, "all_bom_union.xlsx")
BOM_SHEET     = "BOM 2016-2026"
DISPATCH_PATH = os.path.join(SRC_DIR, "Dispatch_Board_Master_2019-2025.csv")
BABY_PATH     = os.path.join(SRC_DIR, "All Shipping Data BABY.xlsm")
BABY_SHEET    = "Master List"
JCM_PATH      = os.path.join(SRC_DIR, "Job Code Master List.xlsx")
JCM_SHEET     = "Master List"

# ── Plant normalization ───────────────────────────────────────────────────────

PLANT_MAP = {
    "sulphur springs": "SS",
    "sulfur springs":  "SS",
    "boulder city":    "BC",
    "plant city":      "PC",
    "ss":              "SS",
    "bc":              "BC",
    "pc":              "PC",
}

def normalize_plant(raw: str) -> str:
    if not raw:
        return ""
    key = raw.strip().lower()
    return PLANT_MAP.get(key, raw.strip().upper()[:2] if len(raw.strip()) >= 2 else raw.strip())


# ── Helpers ───────────────────────────────────────────────────────────────────

def safe(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("none", "nan") else s


def normalize_code(raw) -> str:
    if not raw:
        return ""
    c = str(raw).strip().upper()
    return c if (len(c) == 3 and c.isalpha()) else ""


def code_to_int(code: str) -> int:
    n = 0
    for ch in code.upper():
        n = n * 26 + (ord(ch) - ord("A"))
    return n


def parse_date(value) -> str:
    """Convert Excel datetime object, date, or string to YYYY-MM-DD."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d-%b-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Excel integer serial
    try:
        n = int(float(s))
        if 30000 < n < 60000:
            from openpyxl.utils.datetime import from_excel as _from_excel
            return _from_excel(n).strftime("%Y-%m-%d")
    except Exception:
        pass
    return s


def header_map(row) -> dict:
    """Build {stripped_lower_col_name: index} from a header row."""
    return {str(v).strip().lower(): i for i, v in enumerate(row) if v is not None}


def get_col(hmap: dict, *names):
    """Return first matching column index (case-insensitive) or None."""
    for name in names:
        idx = hmap.get(name.lower())
        if idx is not None:
            return idx
    return None


# ── Data containers ───────────────────────────────────────────────────────────

class JobRecord:
    __slots__ = [
        "bom_project_name", "dispatch_job_name", "shipping_customer",
        "state", "city", "county", "plants",
        "year_released", "date_released",
        "in_bom", "in_dispatch", "in_shipping",
        "bom_row_count", "dispatch_row_count", "shipping_row_count",
        "bom_structures", "dispatch_structures",
        "bom_state", "bom_city",
        "dispatch_state", "dispatch_city",
        "ship_state", "ship_city", "ship_county",
    ]

    def __init__(self):
        self.bom_project_name   = ""
        self.dispatch_job_name  = ""
        self.shipping_customer  = ""
        self.state              = ""
        self.city               = ""
        self.county             = ""
        self.plants             = set()
        self.year_released      = ""
        self.date_released      = ""
        self.in_bom             = False
        self.in_dispatch        = False
        self.in_shipping        = False
        self.bom_row_count      = 0
        self.dispatch_row_count = 0
        self.shipping_row_count = 0
        self.bom_structures     = set()
        self.dispatch_structures = set()
        self.bom_state          = ""
        self.bom_city           = ""
        self.dispatch_state     = ""
        self.dispatch_city      = ""
        self.ship_state         = ""
        self.ship_city          = ""
        self.ship_county        = ""


registry: dict[str, JobRecord] = {}


def get_or_create(code: str) -> JobRecord:
    if code not in registry:
        registry[code] = JobRecord()
    return registry[code]


# ── 1. Load BOM ───────────────────────────────────────────────────────────────

print("=" * 60)
print("PHASE 1 — Job Code Registry Builder")
print("=" * 60)
print()
print("[1/4] Loading BOM (all_bom_union.xlsx)...")
wb_bom = openpyxl.load_workbook(BOM_PATH, read_only=True)
ws_bom = wb_bom[BOM_SHEET]

bom_rows_read  = 0
header_loaded  = False
jc_idx = pname_idx = loc_idx = yr_idx = date_idx = None
struct_idx = state_idx = city_idx = None

for row in ws_bom.iter_rows(values_only=True):
    if not header_loaded:
        hmap = header_map(row)
        all_cols = list(hmap.keys())
        print(f"  Columns found ({len(all_cols)}): {all_cols}")
        jc_idx     = get_col(hmap, "job code", "jobcode", "job_code", "code")
        pname_idx  = get_col(hmap, "project name", "project", "job name")
        loc_idx    = get_col(hmap, "job location", "location", "job_location")
        yr_idx     = get_col(hmap, "year release", "year released", "year", "yr")
        date_idx   = get_col(hmap, "bom release date", "release date", "bom_release_date", "date released", "date")
        struct_idx = get_col(hmap, "structure name", "structure", "struct_name", "structure_name")
        state_idx  = get_col(hmap, "state", "job state")
        city_idx   = get_col(hmap, "city", "job city")
        print(f"  Mapped: job_code={jc_idx}, project_name={pname_idx}, location={loc_idx}, "
              f"year={yr_idx}, date={date_idx}, structure={struct_idx}, state={state_idx}, city={city_idx}")
        if jc_idx is None:
            print("  !! WARNING: No job code column found. Tried: job code, jobcode, job_code, code")
            print("  !! Available columns:", all_cols)
        header_loaded = True
        continue

    if jc_idx is None:
        break

    code = normalize_code(row[jc_idx])
    if not code:
        continue

    bom_rows_read += 1
    rec = get_or_create(code)
    rec.in_bom = True
    rec.bom_row_count += 1

    if pname_idx is not None and not rec.bom_project_name:
        rec.bom_project_name = safe(row[pname_idx])

    if yr_idx is not None:
        yr = safe(row[yr_idx])
        if yr and (not rec.year_released or yr < rec.year_released):
            rec.year_released = yr

    if date_idx is not None:
        d = parse_date(row[date_idx])
        if d and (not rec.date_released or d < rec.date_released):
            rec.date_released = d

    if struct_idx is not None:
        sname = safe(row[struct_idx])
        if sname:
            rec.bom_structures.add(sname)

    if state_idx is not None:
        s = safe(row[state_idx])
        if s and not rec.bom_state:
            rec.bom_state = s

    if city_idx is not None:
        c = safe(row[city_idx])
        if c and not rec.bom_city:
            rec.bom_city = c

    # Fallback: parse "City, ST" from Job Location
    if loc_idx is not None and (not rec.bom_state or not rec.bom_city):
        loc = safe(row[loc_idx])
        if loc and "," in loc:
            parts = [p.strip() for p in loc.split(",", 1)]
            if not rec.bom_city and parts[0]:
                rec.bom_city = parts[0]
            if not rec.bom_state and len(parts) > 1 and parts[1]:
                rec.bom_state = parts[1]

wb_bom.close()
bom_codes = sum(1 for r in registry.values() if r.in_bom)
print(f"  {bom_rows_read:,} rows read  |  {bom_codes:,} unique job codes")


# ── 2. Load Dispatch Board ────────────────────────────────────────────────────

print()
print("[2/4] Loading Dispatch Board (Dispatch_Board_Master_2019-2025.csv)...")
dispatch_rows_read = 0

with open(DISPATCH_PATH, newline="", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    raw_headers = reader.fieldnames or []
    print(f"  Columns found ({len(raw_headers)}): {raw_headers}")

    # Build case-insensitive field lookup
    field_lower = {h.strip().lower(): h for h in raw_headers}

    def get_field(row, *names) -> str:
        for name in names:
            key = field_lower.get(name.lower())
            if key is not None:
                v = row.get(key, "").strip()
                if v:
                    return v
        return ""

    for row in reader:
        code_raw = get_field(row, "job code", "job_code", "jobcode", "code")
        code = normalize_code(code_raw)
        if not code:
            continue

        dispatch_rows_read += 1
        rec = get_or_create(code)
        rec.in_dispatch = True
        rec.dispatch_row_count += 1

        if not rec.dispatch_job_name:
            rec.dispatch_job_name = get_field(row, "job_name", "job name", "jobname", "project name", "project")

        state = get_field(row, "state", "ship_state", "shipping state", "job state")
        city  = get_field(row, "city",  "ship_city",  "shipping city",  "job city")
        if state and not rec.dispatch_state:
            rec.dispatch_state = state
        if city and not rec.dispatch_city:
            rec.dispatch_city = city

        plant = normalize_plant(get_field(row, "plant", "plant name", "plant_name", "mfg", "location"))
        if plant:
            rec.plants.add(plant)

        struct_id = get_field(row, "structure_id", "structure id", "structure name",
                              "structure", "struct_id", "id")
        if struct_id:
            rec.dispatch_structures.add(struct_id)

dispatch_codes = sum(1 for r in registry.values() if r.in_dispatch)
print(f"  {dispatch_rows_read:,} rows read  |  {dispatch_codes:,} unique job codes")


# ── 3. Load ERP Shipping (Master List) ───────────────────────────────────────

print()
print("[3/4] Loading ERP Shipping (All Shipping Data BABY.xlsm -> Master List)...")
wb_baby = openpyxl.load_workbook(BABY_PATH, read_only=True, keep_vba=True)

# List available sheets so we can debug if Master List is named differently
available_sheets = wb_baby.sheetnames
print(f"  Available sheets: {available_sheets}")
if BABY_SHEET not in available_sheets:
    print(f"  !! Sheet '{BABY_SHEET}' not found. Available: {available_sheets}")
    ws_baby = wb_baby.active
    print(f"  Falling back to active sheet: {ws_baby.title}")
else:
    ws_baby = wb_baby[BABY_SHEET]

ship_rows_read = 0
ship_header_loaded = False
sjc_idx = scust_idx = sstate_idx = scity_idx = scounty_idx = splant_idx = None

for row in ws_baby.iter_rows(values_only=True):
    if not ship_header_loaded:
        hmap_s = header_map(row)
        all_scols = list(hmap_s.keys())
        print(f"  Columns found ({len(all_scols)}): {all_scols[:20]}{'...' if len(all_scols) > 20 else ''}")
        sjc_idx     = get_col(hmap_s, "job code", "job code ", "jobcode", "job_code", "code")
        scust_idx   = get_col(hmap_s, "invoiced customer", "invoiced custumer", "customer", "bill to", "billed to")
        sstate_idx  = get_col(hmap_s, "shippings state", "shipping state", "ship state", "state")
        scity_idx   = get_col(hmap_s, "shipping city", "ship city", "city")
        scounty_idx = get_col(hmap_s, "shipping county", "county")
        splant_idx  = get_col(hmap_s, "plant", "plant name", "mfg plant", "mfg")
        print(f"  Mapped: job_code={sjc_idx}, customer={scust_idx}, state={sstate_idx}, "
              f"city={scity_idx}, county={scounty_idx}, plant={splant_idx}")
        if sjc_idx is None:
            print("  !! WARNING: No job code column found in Shipping!")
            print("  !! Available columns:", all_scols)
        ship_header_loaded = True
        continue

    if sjc_idx is None:
        break

    code = normalize_code(row[sjc_idx])
    if not code:
        continue

    ship_rows_read += 1
    rec = get_or_create(code)
    rec.in_shipping = True
    rec.shipping_row_count += 1

    if scust_idx is not None and not rec.shipping_customer:
        rec.shipping_customer = safe(row[scust_idx])

    if sstate_idx is not None and not rec.ship_state:
        rec.ship_state = safe(row[sstate_idx])
    if scity_idx is not None and not rec.ship_city:
        rec.ship_city = safe(row[scity_idx])
    if scounty_idx is not None and not rec.ship_county:
        rec.ship_county = safe(row[scounty_idx])

    if splant_idx is not None:
        p = normalize_plant(safe(row[splant_idx]))
        if p:
            rec.plants.add(p)

wb_baby.close()
ship_codes = sum(1 for r in registry.values() if r.in_shipping)
print(f"  {ship_rows_read:,} rows read  |  {ship_codes:,} unique job codes")


# ── 4. Load Job Code Master List (universe) ───────────────────────────────────

print()
print("[4/4] Loading Job Code Master List (universe)...")
wb_jcm = openpyxl.load_workbook(JCM_PATH, read_only=True)
jcm_sheets = wb_jcm.sheetnames
print(f"  Available sheets: {jcm_sheets}")
ws_jcm = wb_jcm[JCM_SHEET] if JCM_SHEET in jcm_sheets else wb_jcm.active

jcm_hmap_loaded = False
jcm_jc_idx = None
jcm_codes_added = 0

for row in ws_jcm.iter_rows(values_only=True):
    if not jcm_hmap_loaded:
        jcm_hmap = header_map(row)
        print(f"  Columns: {list(jcm_hmap.keys())}")
        jcm_jc_idx = get_col(jcm_hmap, "job code", "code", "jobcode")
        jcm_hmap_loaded = True
        continue
    if jcm_jc_idx is None:
        break
    code = normalize_code(row[jcm_jc_idx])
    if code and code not in registry:
        registry[code] = JobRecord()
        jcm_codes_added += 1

wb_jcm.close()
print(f"  {jcm_codes_added} additional codes added from Job Code Master List")


# ── 5. Resolve location priority ─────────────────────────────────────────────

for rec in registry.values():
    rec.state  = rec.bom_state  or rec.dispatch_state  or rec.ship_state
    rec.city   = rec.bom_city   or rec.dispatch_city   or rec.ship_city
    rec.county = rec.ship_county


# ── 6. Write Job_Code_Registry.xlsx ──────────────────────────────────────────

print()
print(f"Writing Job_Code_Registry.xlsx ({len(registry):,} job codes)...")

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EEF2F7")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
thin         = Side(style="thin", color="CCCCCC")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMNS = [
    ("Job Code",                    10,  "center"),
    ("BOM Project Name",            44,  "left"),
    ("Dispatch Job Name",           38,  "left"),
    ("Shipping Customer",           34,  "left"),
    ("State",                       14,  "left"),
    ("City",                        22,  "left"),
    ("County",                      22,  "left"),
    ("Plant(s)",                    14,  "center"),
    ("Year Released",               14,  "center"),
    ("Date Released to Production", 26,  "center"),
    ("In BOM?",                     10,  "center"),
    ("In Dispatch?",                13,  "center"),
    ("In Shipping?",                13,  "center"),
    ("BOM Row Count",               15,  "center"),
    ("Dispatch Row Count",          17,  "center"),
    ("Shipping Row Count",          17,  "center"),
    ("Sources",                     32,  "left"),
]

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Registry"

for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
    cell = ws_out.cell(row=1, column=col_idx, value=header)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = ALIGN_CENTER
    cell.border    = BORDER
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

ws_out.row_dimensions[1].height = 18
ws_out.freeze_panes = "A2"

sorted_codes = sorted(registry.keys(), key=code_to_int)

for row_idx, code in enumerate(sorted_codes, start=2):
    rec = registry[code]
    alt = (row_idx % 2 == 0)
    fill = ALT_FILL if alt else None

    sources_parts = []
    if rec.in_bom:      sources_parts.append("BOM")
    if rec.in_dispatch: sources_parts.append("Dispatch")
    if rec.in_shipping: sources_parts.append("Shipping")
    sources_str = ", ".join(sources_parts)

    plants_str = ", ".join(sorted(rec.plants)) if rec.plants else ""

    row_values = [
        code,
        rec.bom_project_name,
        rec.dispatch_job_name,
        rec.shipping_customer,
        rec.state,
        rec.city,
        rec.county,
        plants_str,
        rec.year_released,
        rec.date_released,
        "Y" if rec.in_bom      else "N",
        "Y" if rec.in_dispatch else "N",
        "Y" if rec.in_shipping else "N",
        rec.bom_row_count      if rec.bom_row_count      else "",
        rec.dispatch_row_count if rec.dispatch_row_count else "",
        rec.shipping_row_count if rec.shipping_row_count else "",
        sources_str,
    ]

    aligns = [c[2] for c in COLUMNS]

    for col_idx, (val, align) in enumerate(zip(row_values, aligns), start=1):
        cell = ws_out.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = BODY_FONT
        cell.alignment = ALIGN_CENTER if align == "center" else ALIGN_LEFT
        cell.border    = BORDER
        if fill:
            cell.fill = fill

ws_out.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
wb_out.save(OUTPUT_PATH)
print(f"  Saved: {OUTPUT_PATH}")


# ── 7. Update Job Code markdown files ─────────────────────────────────────────

print()
print("Updating Job Code markdown files...")

updated_md = 0
created_md = 0
skipped_md = 0

# Build a lookup: code -> existing filepath (files may be named "AAA.md" or "AAA - Name.md")
existing_files: dict[str, str] = {}
if os.path.isdir(JC_DIR):
    for fname in os.listdir(JC_DIR):
        if not fname.endswith(".md"):
            continue
        stem = os.path.splitext(fname)[0]
        # Extract code: first 3 chars if alpha
        code_part = stem[:3].upper()
        if len(code_part) == 3 and code_part.isalpha():
            if code_part not in existing_files:
                existing_files[code_part] = os.path.join(JC_DIR, fname)

for code in sorted_codes:
    rec = registry[code]
    if not rec.bom_structures and not rec.dispatch_structures:
        continue

    bom_lines  = sorted(rec.bom_structures)
    disp_lines = sorted(rec.dispatch_structures)

    registry_block = "\n## Structure Registry\n\n"
    if bom_lines:
        registry_block += "### BOM Structures\n"
        registry_block += "\n".join(f"- {s}" for s in bom_lines)
        registry_block += "\n\n"
    if disp_lines:
        registry_block += "### Dispatch Structures\n"
        registry_block += "\n".join(f"- {s}" for s in disp_lines)
        registry_block += "\n"

    filepath = existing_files.get(code)
    if filepath:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
        if "## Structure Registry" in content:
            skipped_md += 1
            continue
        with open(filepath, "a", encoding="utf-8") as f:
            f.write("\n" + registry_block)
        updated_md += 1
    else:
        new_path = os.path.join(JC_DIR, f"{code}.md")
        with open(new_path, "w", encoding="utf-8") as f:
            f.write(f"---\njob_code: {code}\n---\n\n# {code}\n{registry_block}")
        created_md += 1

print(f"  Updated  (appended registry): {updated_md:,}")
print(f"  Created  (new files):         {created_md:,}")
print(f"  Skipped  (already had section): {skipped_md:,}")


# ── Summary ───────────────────────────────────────────────────────────────────

in_bom      = sum(1 for r in registry.values() if r.in_bom)
in_dispatch = sum(1 for r in registry.values() if r.in_dispatch)
in_shipping = sum(1 for r in registry.values() if r.in_shipping)
in_all3     = sum(1 for r in registry.values() if r.in_bom and r.in_dispatch and r.in_shipping)
bom_only    = sum(1 for r in registry.values() if r.in_bom and not r.in_dispatch and not r.in_shipping)
ship_no_bom = sum(1 for r in registry.values() if r.in_shipping and not r.in_bom)
no_source   = sum(1 for r in registry.values() if not r.in_bom and not r.in_dispatch and not r.in_shipping)

print()
print("=" * 60)
print("REGISTRY SUMMARY")
print("=" * 60)
print(f"Total job codes:                 {len(registry):,}")
print(f"In BOM:                          {in_bom:,}")
print(f"In Dispatch:                     {in_dispatch:,}")
print(f"In Shipping:                     {in_shipping:,}")
print(f"In all three sources:            {in_all3:,}   (target ~1,483)")
print(f"BOM only (no dispatch/shipping): {bom_only:,}   (target ~230)")
print(f"Shipping but no BOM:             {ship_no_bom:,}  (target ~618)")
print(f"In Master List only (no sources):{no_source:,}   (target ~464)")
print()
print(f"Output Excel:  {OUTPUT_PATH}")
print(f"Markdown files updated: {updated_md + created_md}")
print()
print("Done.")
