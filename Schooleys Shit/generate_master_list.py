"""
Generates Job Code Master List.xlsx from All Shipping Data BABY.xlsm.
Columns: Year, Date Released to Production, Job Code, Job Name,
         State, City, County, BOM Count, Shipping Count
"""

import os
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────

EXCEL_PATH = (
    r"C:\Users\AlecSchooley\Desktop\Schooley Chaotic mind"
    r"\Projects\Shipping Data\All Shipping Data BABY.xlsm"
)
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "Job Code Master List.xlsx")
SHEET_NAME  = "Job Code Data"

# ── Helpers ───────────────────────────────────────────────────────────────────

def code_to_int(code: str) -> int:
    n = 0
    for ch in code.upper():
        n = n * 26 + (ord(ch) - ord("A"))
    return n


def int_to_code(n: int) -> str:
    chars = []
    for _ in range(3):
        chars.append(chr(n % 26 + ord("A")))
        n //= 26
    return "".join(reversed(chars))


def non_empty(row, limit: int) -> int:
    return sum(1 for v in row[:limit] if v is not None and str(v).strip())


def safe(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("none", "nan") else s


def title_case(s: str) -> str:
    if not s:
        return ""
    return s.strip().title()


def extract_hyperlink_text(value) -> str:
    if not value:
        return ""
    s = str(value)
    m = re.search(r'HYPERLINK\("[^"]*",\s*"([^"]+)"\)', s)
    return m.group(1).strip() if m else s.strip()


# ── Load Excel ────────────────────────────────────────────────────────────────

print("Reading source data...")
wb_src = openpyxl.load_workbook(EXCEL_PATH, read_only=True, keep_vba=True)
ws_src = wb_src[SHEET_NAME]

HDR = {
    "year": 0, "code": 1, "project": 2, "county": 3, "zip": 4,
    "city": 5, "state": 6, "customer": 7,
    "plant": 9, "bom_pdf": 10,
}

raw_rows = []
for i, row in enumerate(ws_src.iter_rows(values_only=True)):
    if i == 0:
        continue
    code_raw = row[HDR["code"]]
    if not code_raw or not isinstance(code_raw, str):
        continue
    code = code_raw.strip().upper()
    if not (len(code) == 3 and code.isalpha()):
        continue

    filled = non_empty(row, 12)
    raw_rows.append({
        "_filled": filled,
        "code":    code,
        "year":    row[HDR["year"]],
        "project": row[HDR["project"]],
        "county":  row[HDR["county"]],
        "city":    row[HDR["city"]],
        "state":   row[HDR["state"]],
    })

wb_src.close()

# Deduplicate: prefer most-filled row, tiebreak on higher year
lookup: dict[str, dict] = {}
for r in raw_rows:
    code = r["code"]
    existing = lookup.get(code)
    def yr(x):
        try:
            return int(x["year"]) if x["year"] else 0
        except Exception:
            return 0
    if existing is None or (r["_filled"], yr(r)) > (existing["_filled"], yr(existing)):
        lookup[code] = r

# Build full sequence AAA → max
SEQ_START = "AAA"
if lookup:
    max_code = max(lookup.keys(), key=code_to_int)
    all_codes = [int_to_code(i) for i in range(code_to_int(SEQ_START), code_to_int(max_code) + 1)]
else:
    all_codes = []

print(f"  {len(raw_rows)} rows, {len(lookup)} unique codes, {len(all_codes)} total in sequence")

# ── Build output workbook ─────────────────────────────────────────────────────

wb_out = Workbook()
ws = wb_out.active
ws.title = "Master List"

# Header style
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EEF2F7")   # light blue-grey
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMNS = [
    ("Year",                       10),
    ("Date Released to Production", 22),
    ("Job Code",                    10),
    ("Job Name",                    42),
    ("State",                       18),
    ("City",                        20),
    ("County",                      22),
    ("BOM Count",                   12),
    ("Shipping Count",              15),
]

# Write headers
for col_idx, (header, width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font    = HEADER_FONT
    cell.fill    = HEADER_FILL
    cell.alignment = ALIGN_CENTER
    cell.border  = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 18
ws.freeze_panes = "A2"

# Write data rows
for row_idx, code in enumerate(all_codes, start=2):
    data = lookup.get(code)

    if data:
        year     = safe(data["year"])
        try:
            year = str(int(float(year))) if year else ""
        except Exception:
            pass
        job_name = safe(data["project"])
        state    = title_case(safe(data["state"]))
        city     = title_case(safe(data["city"]))
        county   = title_case(safe(data["county"]))
    else:
        year = job_name = state = city = county = ""

    alt = (row_idx % 2 == 0)
    fill = ALT_FILL if alt else None

    values = [year, "", code, job_name, state, city, county, "", ""]
    aligns = [ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER, ALIGN_LEFT,
              ALIGN_LEFT, ALIGN_LEFT, ALIGN_LEFT, ALIGN_CENTER, ALIGN_CENTER]

    for col_idx, (val, align) in enumerate(zip(values, aligns), start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = BODY_FONT
        cell.alignment = align
        cell.border    = BORDER
        if fill:
            cell.fill = fill

ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

# ── Save ──────────────────────────────────────────────────────────────────────

wb_out.save(OUTPUT_PATH)
print(f"\nSaved: {OUTPUT_PATH}")
print(f"  Rows written: {len(all_codes)}")
print(f"  Populated:    {sum(1 for c in all_codes if c in lookup)}")
print(f"  Placeholders: {sum(1 for c in all_codes if c not in lookup)}")
