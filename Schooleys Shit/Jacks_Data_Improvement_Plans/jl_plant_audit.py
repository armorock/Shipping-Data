"""
Reads output/jobcode_db.json and produces two Excel reports:

  output/plant_conflict_report.xlsx
    All job codes where plant data doesn't fully agree across sources.
    Includes CONFLICT, CONSENSUS_OVERRIDE, and cases where Alec's markdown
    plant differs from the data-derived plant.

  output/problem_children.xlsx
    The short list requiring manual review: CONFLICT and CONSENSUS_OVERRIDE
    rows only, sorted by severity (CONFLICT first).

Run: python jl_plant_audit.py  (after running jl_build_jobcode_db.py first)
"""

import os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE       = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE, "output")
DB_PATH    = os.path.join(OUTPUT_DIR, "jobcode_db.json")

# Colours
RED    = "00FF0000"
ORANGE = "00FF8C00"
YELLOW = "00FFD700"
GREEN  = "0000AA00"
GREY   = "00AAAAAA"
HEADER_FILL = PatternFill("solid", fgColor="00404040")
HEADER_FONT = Font(color="00FFFFFF", bold=True)

RESOLUTION_COLOURS = {
    "CONFLICT":           PatternFill("solid", fgColor=RED),
    "CONSENSUS_OVERRIDE": PatternFill("solid", fgColor=ORANGE),
    "MATCH":              PatternFill("solid", fgColor=GREEN),
    "SINGLE_SOURCE":      PatternFill("solid", fgColor="00DDDDDD"),
    "NO_DATA":            PatternFill("solid", fgColor=GREY),
}


def load_db():
    with open(DB_PATH, encoding="utf-8") as f:
        return json.load(f)


def header_row(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def autofit(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 50)


def write_conflict_report(records):
    COLS = [
        "Job Code", "Project Name", "State", "City",
        "Alec's Plant (markdown)", "Data Plant", "Plant Source",
        "Resolution",
        "Dispatch Plant", "Shipping Plant", "Markdown Plant",
        "Conflict Detail",
        "In BOM?", "In Dispatch?", "In Shipping?",
        "BOM Rows", "Dispatch Rows", "Shipping Rows",
    ]

    # Include rows where resolution is not clean, OR where alec's plant
    # differs from the derived plant (catches cases where one source is missing)
    def should_include(r):
        res = r.get("plant_resolution", "")
        if res in ("CONFLICT", "CONSENSUS_OVERRIDE"):
            return True
        # Also flag where alec wrote a plant but it differs from data
        alec = r.get("plant_alec") or ""
        data = r.get("plant") or ""
        if alec and data and alec != data:
            return True
        return False

    rows = [r for r in records if should_include(r)]
    rows.sort(key=lambda r: (
        0 if r.get("plant_resolution") == "CONFLICT" else
        1 if r.get("plant_resolution") == "CONSENSUS_OVERRIDE" else 2,
        r["job_code"]
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plant Conflicts"
    ws.freeze_panes = "A2"
    header_row(ws, COLS)

    for r in rows:
        resolution = r.get("plant_resolution", "")
        row_data = [
            r["job_code"],
            r.get("project_name") or "",
            r.get("shipping_state") or "",
            r.get("shipping_city") or "",
            r.get("plant_alec") or "",
            r.get("plant") or "",
            r.get("plant_source") or "",
            resolution,
            "",  # Dispatch plant — extracted below from conflict detail
            "",  # Shipping plant
            r.get("plant_alec") or "",
            r.get("plant_conflict") or "",
            "Y" if r.get("in_bom") else "N",
            "Y" if r.get("in_dispatch") else "N",
            "Y" if r.get("in_shipping") else "N",
            r.get("bom_row_count", 0),
            r.get("dispatch_row_count", 0),
            r.get("shipping_row_count", 0),
        ]
        ws.append(row_data)

        # Colour the Resolution cell (col 8)
        res_cell = ws.cell(ws.max_row, 8)
        fill = RESOLUTION_COLOURS.get(resolution)
        if fill:
            res_cell.fill = fill

        # Bold the job code
        ws.cell(ws.max_row, 1).font = Font(bold=True)

    autofit(ws)
    ws.auto_filter.ref = ws.dimensions

    out = os.path.join(OUTPUT_DIR, "plant_conflict_report.xlsx")
    wb.save(out)
    return out, len(rows)


def write_problem_children(records):
    COLS = [
        "Job Code", "Project Name", "State", "City",
        "Alec's Plant", "Data Plant", "Plant Source",
        "Resolution", "Conflict Detail",
        "Independent Sources", "Action Needed",
    ]

    rows = [r for r in records
            if r.get("plant_resolution") in ("CONFLICT", "CONSENSUS_OVERRIDE")]
    rows.sort(key=lambda r: (
        0 if r.get("plant_resolution") == "CONFLICT" else 1,
        r["job_code"]
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Problem Children"
    ws.freeze_panes = "A2"
    header_row(ws, COLS)

    for r in rows:
        resolution = r.get("plant_resolution", "")
        conflict   = r.get("plant_conflict") or ""

        # Count independent sources mentioned in plant_source
        src = r.get("plant_source") or ""
        independent_count = len([s for s in src.split("+") if s.strip()])

        action = (
            "Verify: BOM plant differs from 3-source consensus" if resolution == "CONSENSUS_OVERRIDE"
            else "Verify: sources disagree, no clear consensus"
        )

        row_data = [
            r["job_code"],
            r.get("project_name") or "",
            r.get("shipping_state") or "",
            r.get("shipping_city") or "",
            r.get("plant_alec") or "",
            r.get("plant") or "",
            src,
            resolution,
            conflict,
            independent_count,
            action,
        ]
        ws.append(row_data)

        res_cell = ws.cell(ws.max_row, 8)
        fill = RESOLUTION_COLOURS.get(resolution)
        if fill:
            res_cell.fill = fill

        ws.cell(ws.max_row, 1).font = Font(bold=True)

    autofit(ws)
    ws.auto_filter.ref = ws.dimensions

    out = os.path.join(OUTPUT_DIR, "problem_children.xlsx")
    wb.save(out)
    return out, len(rows)


def main():
    if not os.path.exists(DB_PATH):
        print(f"ERROR: {DB_PATH} not found. Run jl_build_jobcode_db.py first.")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    records = load_db()
    print(f"Loaded {len(records):,} records from jobcode_db.json")

    path1, count1 = write_conflict_report(records)
    print(f"  plant_conflict_report.xlsx: {count1:,} rows -> {path1}")

    path2, count2 = write_problem_children(records)
    print(f"  problem_children.xlsx:      {count2:,} rows -> {path2}")

    if count2 == 0:
        print("\nNo problem children found — all plant data agrees or has only one source.")
    elif count2 <= 20:
        print(f"\n{count2} problem children — small list, easy to review manually.")
    else:
        print(f"\n{count2} problem children — if this seems high, check the conflict")
        print("resolution logic in jl_build_jobcode_db.py for contamination issues.")


if __name__ == "__main__":
    main()
