"""
Builds output/jobcode_db.json from all four data sources:
  - BOM union (all_bom_union.xlsx)
  - Dispatch Board (Dispatch_Board_Master_2019-2025.csv)
  - ERP Shipping (All Shipping Data BABY.xlsm)
  - Job Code markdown files (Job Codes/*.md)

Every field in the output carries a _source tag. Conflicts are stored inline
with a _resolution label: MATCH, SINGLE_SOURCE, CONFLICT, or CONSENSUS_OVERRIDE.

CONSENSUS_OVERRIDE fires when BOM says X but all three independent non-BOM
sources (Dispatch, Shipping, Markdown) agree on Y. Those rows land in
problem_children.xlsx for manual review.

Run: python jl_build_jobcode_db.py
"""

import os, re, json, csv
from datetime import datetime, date
import openpyxl

BASE        = os.path.dirname(os.path.abspath(__file__))
MASTER      = os.path.join(BASE, "..", "MASTER CSV FILES")
JOBCODES_DIR = os.path.join(BASE, "..", "Job Codes")
OUTPUT_DIR  = os.path.join(BASE, "output")

BOM_PATH      = os.path.join(MASTER, "all_bom_union.xlsx")
BOM_SHEET     = "BOM 2016-2026"
DISPATCH_PATH = os.path.join(MASTER, "Dispatch_Board_Master_2019-2025.csv")
BABY_PATH     = os.path.join(MASTER, "All Shipping Data BABY.xlsm")
BABY_SHEET    = "Master List"

PLANT_MAP = {
    "sulphur springs": "SS", "sulfur springs": "SS",
    "boulder city": "BC",    "plant city": "PC",
    "ss": "SS", "bc": "BC", "pc": "PC",
}

STATE_ABBREVS = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY", "district of columbia": "DC",
}

CONFIDENCE = {"BOM": 0, "Dispatch": 1, "Shipping": 2, "Markdown": 3}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def hdr(row):
    return {str(v).strip().lower(): i for i, v in enumerate(row) if v is not None}

def col(hmap, *keys):
    for k in keys:
        if k.lower() in hmap:
            return hmap[k.lower()]
    return None

def safe(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.startswith("=") else s

def normalize_code(val):
    if not val:
        return None
    v = re.sub(r"\s+", "", str(val)).upper()
    return v if re.match(r"^[A-E][A-Z0-9]{2}$", v) else None

def normalize_plant(val):
    if not val:
        return None
    raw = str(val).strip()
    # Handle multi-plant like "SS/BC"
    if "/" in raw:
        parts = [normalize_plant(p.strip()) for p in raw.split("/")]
        parts = [p for p in parts if p]
        return "/".join(sorted(set(parts))) if parts else None
    mapped = PLANT_MAP.get(raw.lower())
    if mapped:
        return mapped
    upper = raw.upper()
    return upper if upper in ("SS", "BC", "PC") else None

def to_state_abbrev(raw):
    if not raw:
        return None
    raw = raw.strip()
    if len(raw) == 2:
        return raw.upper()
    return STATE_ABBREVS.get(raw.lower())

def parse_location(loc_str):
    """'City, ST' or 'City, State' -> (city, state_abbrev)"""
    if not loc_str:
        return None, None
    loc = re.sub(r",?\s*(USA|US)$", "", loc_str.strip(), flags=re.IGNORECASE).strip()
    if "," in loc:
        city, _, state_raw = loc.rpartition(",")
        return city.strip() or None, to_state_abbrev(state_raw.strip())
    return loc or None, None

def to_year(val):
    if not val:
        return None
    try:
        return int(str(val).strip()[:4])
    except (ValueError, TypeError):
        return None

def to_date_str(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None

def parse_yaml_frontmatter(filepath):
    fields = {}
    try:
        with open(filepath, encoding="utf-8", errors="replace") as f:
            content = f.read()
        m = re.match(r"^---\s*\n(.*?)\n---", content, re.DOTALL)
        if not m:
            return fields
        for line in m.group(1).splitlines():
            if ":" in line:
                key, _, val = line.partition(":")
                key = key.strip().lower()
                val = val.strip().strip("\"'")
                val = re.sub(r"\[\[(.+?)\]\]", r"\1", val)
                fields[key] = val or None
    except Exception:
        pass
    return fields


# ---------------------------------------------------------------------------
# Conflict resolution
# ---------------------------------------------------------------------------

def resolve(sources):
    """
    sources: list of (group_name, value) — group_name in CONFIDENCE dict.
    Filters out empty/None values, then applies the consensus rules.
    Returns (value, source_tag, resolution, conflict_detail).
    """
    available = [(g, v) for g, v in sources if v]
    if not available:
        return None, None, "NO_DATA", None
    if len(available) == 1:
        return available[0][1], available[0][0], "SINGLE_SOURCE", None

    unique_vals = set(v for _, v in available)
    if len(unique_vals) == 1:
        best = min(available, key=lambda x: CONFIDENCE.get(x[0], 99))
        return best[1], best[0], "MATCH", None

    # Conflict — check if BOM has a value
    bom_vals   = [v for g, v in available if g == "BOM"]
    non_bom    = [(g, v) for g, v in available if g != "BOM"]
    non_bom_vals = set(v for _, v in non_bom)

    if bom_vals:
        bom_val = bom_vals[0]

        if len(non_bom_vals) == 1:
            # All non-BOM sources agree on something different from BOM
            consensus = list(non_bom_vals)[0]
            non_bom_groups = {g for g, _ in non_bom}
            has_d = "Dispatch" in non_bom_groups
            has_s = "Shipping" in non_bom_groups
            has_m = "Markdown" in non_bom_groups
            independent = sum([has_d, has_s, has_m])

            if independent >= 3:
                src_str = "+".join(sorted(non_bom_groups))
                detail = f"BOM='{bom_val}'; {src_str} all say '{consensus}'"
                return consensus, src_str, "CONSENSUS_OVERRIDE", detail

        # No clean consensus — keep BOM
        parts = "; ".join(f"{g}='{v}'" for g, v in
                          sorted(available, key=lambda x: CONFIDENCE.get(x[0], 99)))
        return bom_val, "BOM", "CONFLICT", parts

    # No BOM data — use highest-confidence available
    best = min(available, key=lambda x: CONFIDENCE.get(x[0], 99))
    if len(unique_vals) > 1:
        parts = "; ".join(f"{g}='{v}'" for g, v in
                          sorted(available, key=lambda x: CONFIDENCE.get(x[0], 99)))
        return best[1], best[0], "CONFLICT", parts
    return best[1], best[0], "MATCH", None


# ---------------------------------------------------------------------------
# Source readers
# ---------------------------------------------------------------------------

def read_bom():
    print("Reading BOM union...")
    data = {}
    wb = openpyxl.load_workbook(BOM_PATH, read_only=True)
    ws = wb[BOM_SHEET]
    hmap = None
    jc = pn = loc = zp = ctr = yr = dt = st = ct = None

    for row in ws.iter_rows(values_only=True):
        if hmap is None:
            hmap = hdr(row)
            jc  = col(hmap, "job code", "jobcode", "job_code", "code")
            pn  = col(hmap, "project name", "project", "job name")
            loc = col(hmap, "job location", "location", "job_location")
            zp  = col(hmap, "zip code", "zip", "postal code")
            ctr = col(hmap, "contractor", "contractor name")
            yr  = col(hmap, "year release", "year released", "year")
            dt  = col(hmap, "bom release date", "release date", "date released")
            st  = col(hmap, "state", "job state")
            ct  = col(hmap, "city", "job city")
            continue

        code = normalize_code(row[jc] if jc is not None else None)
        if not code:
            continue

        if code not in data:
            data[code] = {
                "project_name": None, "city": None, "state": None,
                "zip": None, "contractor": None,
                "year_released": None, "date_released": None, "row_count": 0,
            }
        rec = data[code]
        rec["row_count"] += 1

        if not rec["project_name"] and pn is not None:
            v = safe(row[pn]); rec["project_name"] = v or None

        # State / city — explicit columns first, then parse from Job Location
        raw_st = safe(row[st]) if st is not None else ""
        raw_ct = safe(row[ct]) if ct is not None else ""
        if not raw_st or not raw_ct:
            loc_str = safe(row[loc]) if loc is not None else ""
            ct2, st2 = parse_location(loc_str)
            if not raw_ct: raw_ct = ct2 or ""
            if not raw_st: raw_st = to_state_abbrev(st2) or st2 or ""
        if raw_ct and not rec["city"]:  rec["city"]  = raw_ct
        if raw_st and not rec["state"]: rec["state"] = raw_st[:2].upper() if len(raw_st) > 2 else raw_st.upper()

        if not rec["zip"] and zp is not None:
            z = safe(row[zp])
            if z: rec["zip"] = z[:5]

        if not rec["contractor"] and ctr is not None:
            c = safe(row[ctr])
            if c: rec["contractor"] = c

        if not rec["year_released"] and yr is not None:
            y = to_year(row[yr])
            if y: rec["year_released"] = y

        if not rec["date_released"] and dt is not None:
            d = to_date_str(row[dt])
            if d: rec["date_released"] = d

    wb.close()
    print(f"  {len(data):,} job codes from BOM")
    return data


def read_dispatch():
    print("Reading Dispatch Board...")
    data = {}
    with open(DISPATCH_PATH, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = normalize_code(row.get("job_code", ""))
            if not code:
                continue
            if code not in data:
                data[code] = {
                    "job_name": None, "city": None, "state": None,
                    "plant": None, "_plant_counts": {}, "row_count": 0,
                }
            rec = data[code]
            rec["row_count"] += 1
            if not rec["job_name"] and row.get("job_name"):
                rec["job_name"] = row["job_name"].strip()
            if not rec["city"] and row.get("city"):
                rec["city"] = row["city"].strip()
            if not rec["state"] and row.get("state"):
                rec["state"] = row["state"].strip().upper()[:2]
            plant = normalize_plant(row.get("plant", ""))
            if plant:
                counts = rec["_plant_counts"]
                counts[plant] = counts.get(plant, 0) + 1
                rec["plant"] = max(counts, key=counts.get)
    print(f"  {len(data):,} job codes from Dispatch")
    return data


def read_shipping():
    print("Reading Shipping (BABY) — this may take a minute...")
    data = {}
    wb = openpyxl.load_workbook(BABY_PATH, read_only=True, data_only=True)
    ws = wb[BABY_SHEET]
    hmap = None
    jc = cust = st = ct = co = zp = pl = None

    for row in ws.iter_rows(values_only=True):
        if hmap is None:
            hmap = hdr(row)
            jc   = col(hmap, "job code", "job code ", "jobcode", "job_code")
            cust = col(hmap, "invoiced customer", "invoiced custumer", "customer", "bill to")
            st   = col(hmap, "shippings state", "shipping state", "ship state", "state")
            ct   = col(hmap, "shipping city", "ship city", "city")
            co   = col(hmap, "shipping county", "county")
            zp   = col(hmap, "shipping zip", "zip", "postal code")
            pl   = col(hmap, "plant", "plant name", "mfg plant", "mfg")
            continue

        code = normalize_code(row[jc] if jc is not None else None)
        if not code:
            continue

        if code not in data:
            data[code] = {
                "customer": None, "city": None, "state": None,
                "county": None, "zip": None, "plant": None,
                "_plant_counts": {}, "row_count": 0,
            }
        rec = data[code]
        rec["row_count"] += 1

        if not rec["customer"] and cust is not None:
            v = safe(row[cust])
            if v: rec["customer"] = v

        if not rec["state"] and st is not None:
            v = safe(row[st])
            if v and len(v) <= 30:
                rec["state"] = v[:2].upper() if len(v) > 2 else v.upper()

        if not rec["city"] and ct is not None:
            v = safe(row[ct])
            if v: rec["city"] = v

        if not rec["county"] and co is not None:
            v = safe(row[co])
            if v: rec["county"] = v

        if not rec["zip"] and zp is not None:
            v = safe(row[zp])
            if v: rec["zip"] = v[:5]

        if pl is not None:
            plant = normalize_plant(row[pl])
            if plant:
                counts = rec["_plant_counts"]
                counts[plant] = counts.get(plant, 0) + 1
                rec["plant"] = max(counts, key=counts.get)

    wb.close()
    print(f"  {len(data):,} job codes from Shipping")
    return data


def read_markdown():
    print("Reading Job Code markdown files...")
    data = {}
    if not os.path.isdir(JOBCODES_DIR):
        print(f"  WARNING: {JOBCODES_DIR} not found, skipping markdown")
        return data

    for fname in os.listdir(JOBCODES_DIR):
        if not fname.endswith(".md"):
            continue
        m = re.match(r"^([A-E][A-Z0-9]{2})", fname.upper())
        if not m:
            continue
        code = m.group(1)
        fields = parse_yaml_frontmatter(os.path.join(JOBCODES_DIR, fname))
        if not fields:
            continue

        raw_state = fields.get("state") or ""
        raw_state = re.sub(r"\[\[(.+?)\]\]", r"\1", raw_state).strip()
        state = to_state_abbrev(raw_state) or (raw_state.upper() if len(raw_state) == 2 else None)

        raw_city = fields.get("city") or ""
        raw_city = re.sub(r"\[\[(.+?)\]\]", r"\1", raw_city).strip()
        city = re.sub(r"\s+[A-Z]{2}$", "", raw_city).strip() or None

        data[code] = {
            "plant":    normalize_plant(fields.get("plant")),
            "customer": fields.get("customer") or None,
            "state":    state,
            "city":     city,
            "zip":      fields.get("zip") or None,
            "year":     to_year(fields.get("year")),
        }

    print(f"  {len(data):,} job codes from markdown files")
    return data


# ---------------------------------------------------------------------------
# Merge
# ---------------------------------------------------------------------------

def merge(bom, disp, ship, md):
    print("Merging and resolving conflicts...")
    all_codes = sorted(
        set(list(bom) + list(disp) + list(ship) + list(md))
    )
    records = []

    for code in all_codes:
        b = bom.get(code, {})
        d = disp.get(code, {})
        s = ship.get(code, {})
        m = md.get(code, {})

        rec = {"job_code": code}

        # Project name
        rec["project_name"] = b.get("project_name") or d.get("job_name") or None
        rec["project_name_source"] = (
            "BOM" if b.get("project_name") else
            "Dispatch" if d.get("job_name") else None
        )

        # Shipping city
        cv, cs, cr, cc = resolve([
            ("BOM",      b.get("city")),
            ("Dispatch", d.get("city")),
            ("Shipping", s.get("city")),
            ("Markdown", m.get("city")),
        ])
        rec.update(shipping_city=cv, shipping_city_source=cs,
                   shipping_city_resolution=cr, shipping_city_conflict=cc)

        # Shipping state
        sv, ss, sr, sc = resolve([
            ("BOM",      b.get("state")),
            ("Dispatch", d.get("state")),
            ("Shipping", s.get("state")),
            ("Markdown", m.get("state")),
        ])
        rec.update(shipping_state=sv, shipping_state_source=ss,
                   shipping_state_resolution=sr, shipping_state_conflict=sc)

        # Zip
        zv, zs, zr, zc = resolve([
            ("BOM",      b.get("zip")),
            ("Shipping", s.get("zip")),
            ("Markdown", m.get("zip")),
        ])
        rec.update(shipping_zip=zv, shipping_zip_source=zs,
                   shipping_zip_resolution=zr, shipping_zip_conflict=zc)

        # County (Shipping only)
        rec["shipping_county"]        = s.get("county")
        rec["shipping_county_source"] = "Shipping" if s.get("county") else None

        # Customer / Contractor
        rec["customer"]        = s.get("customer") or m.get("customer") or None
        rec["customer_source"] = ("Shipping" if s.get("customer") else
                                  "Markdown" if m.get("customer") else None)
        rec["contractor"]        = b.get("contractor") or None
        rec["contractor_source"] = "BOM" if b.get("contractor") else None

        # Plant — BOM typically doesn't carry plant; Dispatch is highest-confidence
        pv, ps, pr, pc = resolve([
            ("Dispatch", d.get("plant")),
            ("Shipping", s.get("plant")),
            ("Markdown", m.get("plant")),
        ])
        rec["plant"]             = pv
        rec["plant_source"]      = ps
        rec["plant_alec"]        = m.get("plant")
        rec["plant_resolution"]  = pr
        rec["plant_conflict"]    = pc

        # Year / date
        rec["year_released"]        = b.get("year_released") or m.get("year") or None
        rec["year_released_source"] = ("BOM" if b.get("year_released") else
                                       "Markdown" if m.get("year") else None)
        rec["date_released"]        = b.get("date_released") or None
        rec["date_released_source"] = "BOM" if b.get("date_released") else None

        # Presence flags
        rec["in_bom"]      = code in bom
        rec["in_dispatch"] = code in disp
        rec["in_shipping"] = code in ship
        rec["in_markdown"] = code in md

        # Row counts
        rec["bom_row_count"]      = b.get("row_count", 0)
        rec["dispatch_row_count"] = d.get("row_count", 0)
        rec["shipping_row_count"] = s.get("row_count", 0)

        records.append(rec)

    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    bom  = read_bom()
    disp = read_dispatch()
    ship = read_shipping()
    md   = read_markdown()

    records = merge(bom, disp, ship, md)

    out = os.path.join(OUTPUT_DIR, "jobcode_db.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, default=str)

    # Summary
    conflicts  = sum(1 for r in records if r.get("plant_resolution") == "CONFLICT")
    overrides  = sum(1 for r in records if r.get("plant_resolution") == "CONSENSUS_OVERRIDE")
    no_data    = sum(1 for r in records if r.get("plant_resolution") == "NO_DATA")
    matches    = sum(1 for r in records if r.get("plant_resolution") == "MATCH")
    single     = sum(1 for r in records if r.get("plant_resolution") == "SINGLE_SOURCE")

    print(f"\nWrote {len(records):,} records -> {out}")
    print(f"  Plant MATCH:              {matches:,}")
    print(f"  Plant SINGLE_SOURCE:      {single:,}")
    print(f"  Plant CONFLICT:           {conflicts:,}  <- needs review")
    print(f"  Plant CONSENSUS_OVERRIDE: {overrides:,}  <- needs review")
    print(f"  Plant NO_DATA:            {no_data:,}")
    print(f"\n  Problem children total: {conflicts + overrides:,}")
    print("\nRun jl_plant_audit.py next to generate the Excel reports.")


if __name__ == "__main__":
    main()
