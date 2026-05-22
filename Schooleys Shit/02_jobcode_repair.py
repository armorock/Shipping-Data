"""
Phase 2 — Job Code Validation & Repair
Uses Phase 1 Job_Code_Registry.xlsx as the valid code universe.
Reads ALL rows from BABY Master List (including nulls/partials).
Outputs 02_repair_output.xlsx (annotated) and jobcode_repair_log.csv.
Never modifies any source file.
"""

import csv
import os
from collections import defaultdict
from datetime import datetime, date, timedelta

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
SRC_DIR         = os.path.join(BASE_DIR, "MASTER CSV FILES")
REGISTRY_PATH   = os.path.join(BASE_DIR, "Job_Code_Registry.xlsx")
BABY_PATH       = os.path.join(SRC_DIR,  "All Shipping Data BABY.xlsm")
BABY_SHEET      = "Master List"
OUTPUT_PATH     = os.path.join(BASE_DIR, "02_repair_output.xlsx")
LOG_PATH        = os.path.join(BASE_DIR, "jobcode_repair_log.csv")

DATE_WINDOW_DAYS = 180

# Rows that shipped before this date cannot have job codes — they predate the system
PRE_JOBCODE_CUTOFF = date(2016, 1, 1)

# Partial codes flagged for manual review (ambiguous, do not auto-assign)
MANUAL_REVIEW_PARTIALS = {"BV", "CJ"}

# ── Helpers ───────────────────────────────────────────────────────────────────

def safe(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("none", "nan") else s


def normalize_code(raw) -> str:
    if not raw:
        return ""
    c = str(raw).strip().upper()
    return c if (len(c) == 3 and c.isalpha()) else ""


def to_date(value):
    """Return a date object or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    try:
        n = int(float(s))
        if 30000 < n < 60000:
            from openpyxl.utils.datetime import from_excel as _fe
            return _fe(n).date()
    except Exception:
        pass
    return None


def cust_match(a: str, b: str) -> bool:
    """True if either string is a substring of the other (case-insensitive, stripped)."""
    if not a or not b:
        return False
    al, bl = a.lower(), b.lower()
    return al in bl or bl in al


def header_map(row) -> dict:
    return {str(v).strip().lower(): i for i, v in enumerate(row) if v is not None}


def get_col(hmap: dict, *names):
    for name in names:
        idx = hmap.get(name.lower())
        if idx is not None:
            return idx
    return None


# ── 1. Load Phase 1 Registry ──────────────────────────────────────────────────

print("=" * 60)
print("PHASE 2 -- Job Code Validation & Repair")
print("=" * 60)
print()
print("[1/3] Loading Phase 1 registry...")

if not os.path.exists(REGISTRY_PATH):
    raise FileNotFoundError(f"Registry not found: {REGISTRY_PATH}\nRun Phase 1 first.")

wb_reg = openpyxl.load_workbook(REGISTRY_PATH, read_only=True)
ws_reg = wb_reg.active

reg_hmap = None
# code -> {customer, state, city, date_released, year_released}
registry: dict[str, dict] = {}

for row in ws_reg.iter_rows(values_only=True):
    if reg_hmap is None:
        reg_hmap = header_map(row)
        continue
    jc_idx    = get_col(reg_hmap, "job code")
    cust_idx  = get_col(reg_hmap, "shipping customer")
    state_idx = get_col(reg_hmap, "state")
    city_idx  = get_col(reg_hmap, "city")
    date_idx  = get_col(reg_hmap, "date released to production")
    yr_idx    = get_col(reg_hmap, "year released")

    code = normalize_code(row[jc_idx] if jc_idx is not None else None)
    if not code:
        continue
    registry[code] = {
        "customer":      safe(row[cust_idx]  if cust_idx  is not None else None),
        "state":         safe(row[state_idx] if state_idx is not None else None).upper(),
        "city":          safe(row[city_idx]  if city_idx  is not None else None).lower(),
        "date_released": to_date(row[date_idx] if date_idx is not None else None),
        "year_released": safe(row[yr_idx]    if yr_idx    is not None else None),
    }

wb_reg.close()
valid_codes = set(registry.keys())
print(f"  {len(valid_codes):,} valid codes loaded from registry")

# ── Build lookup indices ──────────────────────────────────────────────────────

# prefix (1 or 2 chars) -> list of valid 3-char codes
prefix_index: dict[str, list] = defaultdict(list)
for code in valid_codes:
    prefix_index[code[:1]].append(code)
    prefix_index[code[:2]].append(code)

# normalized customer -> set of codes
customer_index: dict[str, set] = defaultdict(set)
for code, rec in registry.items():
    if rec["customer"]:
        customer_index[rec["customer"].lower()].add(code)

# state -> set of codes
state_index: dict[str, set] = defaultdict(set)
for code, rec in registry.items():
    if rec["state"]:
        state_index[rec["state"]].add(code)

print(f"  Prefix index: {len(prefix_index):,} entries")
print(f"  Customer index: {len(customer_index):,} entries")
print(f"  State index: {len(state_index):,} entries")


# ── 2. Scoring engine ─────────────────────────────────────────────────────────

def score_candidate(partial: str, candidate: str,
                    row_cust: str, row_state: str, row_city: str, row_date) -> tuple:
    """
    Returns (score, method_string).
    Criteria (each worth 1 point):
      1. Prefix match (partial codes only)
      2. Customer/contractor match
      3. State match
      4. Date within DATE_WINDOW_DAYS of date_released_to_production
    """
    rec = registry[candidate]
    score = 0
    methods = []

    # 1. Prefix
    if partial and candidate.startswith(partial):
        score += 1
        methods.append("Prefix")

    # 2. Customer
    if cust_match(row_cust, rec["customer"]):
        score += 1
        methods.append("Customer")

    # 3. State
    if row_state and rec["state"] and row_state.upper() == rec["state"]:
        score += 1
        methods.append("State")

    # 4. Date window
    if row_date and rec["date_released"]:
        delta = abs((row_date - rec["date_released"]).days)
        if delta <= DATE_WINDOW_DAYS:
            score += 1
            methods.append("Date")

    method_str = "+".join(methods) if methods else ""
    return score, method_str


def confidence(score: int) -> str:
    if score >= 3: return "High"
    if score == 2: return "Medium"
    if score == 1: return "Low"
    return "n/a"


def find_best_match(partial: str, row_cust: str, row_state: str,
                    row_city: str, row_date, is_null: bool) -> tuple:
    """
    Returns (best_code, score, method, notes) or ("", 0, "", notes).
    partial: the raw job code value from the row (may be blank for nulls).
    is_null: True if the original code was completely blank.
    """
    # Build candidate pool
    if is_null:
        # No prefix — start from customer or state
        candidates = set()
        if row_cust:
            for cust_key, codes in customer_index.items():
                if cust_match(row_cust, cust_key):
                    candidates.update(codes)
        if row_state:
            candidates.update(state_index.get(row_state.upper(), set()))
        # If still empty, nothing to work with
        if not candidates:
            return "", 0, "", "No customer or state match found in registry"
    else:
        # Use prefix as the primary filter
        normalized_partial = str(partial).strip().upper()
        # Check if partial is actually longer than 3 chars and starts with a valid code
        if len(normalized_partial) > 3 and normalized_partial[:3] in valid_codes:
            return normalized_partial[:3], 4, "Truncated(exact prefix)", f"Partial '{partial}' starts with valid code '{normalized_partial[:3]}'"
        candidates = set(prefix_index.get(normalized_partial, []))
        # Also try prefixes of the partial (in case partial is 2 chars of a code)
        if not candidates and len(normalized_partial) >= 2:
            candidates = set(prefix_index.get(normalized_partial[:2], []))
        if not candidates:
            return "", 0, "", f"No registry code starts with '{partial}'"

    # Score all candidates
    scored = []
    for cand in candidates:
        s, method = score_candidate(partial if not is_null else "", cand,
                                    row_cust, row_state, row_city, row_date)
        scored.append((s, cand, method))

    scored.sort(key=lambda x: -x[0])

    if not scored or scored[0][0] == 0:
        return "", 0, "", f"All candidates scored 0 (tried {len(candidates)} codes)"

    best_score, best_code, best_method = scored[0]

    # Build notes
    if len(scored) > 1 and scored[1][0] == best_score:
        # Tie — flag it
        tied = [c for s, c, _ in scored if s == best_score]
        notes = f"Tied {len(tied)} candidates at score {best_score}: {', '.join(tied[:5])}"
        if len(tied) > 5:
            notes += f" (+{len(tied)-5} more)"
    elif len(candidates) == 1 and not is_null:
        notes = f"Single prefix match -> {best_code}"
    else:
        notes = f"Best of {len(candidates)} candidates (score {best_score})"

    return best_code, best_score, best_method, notes


# ── 3. Read ALL BABY Master List rows & annotate ──────────────────────────────

print()
print("[2/3] Reading BABY Master List and annotating...")

wb_baby = openpyxl.load_workbook(BABY_PATH, read_only=True, keep_vba=True)
print(f"  Available sheets: {wb_baby.sheetnames[:8]}...")
ws_baby = wb_baby[BABY_SHEET]

baby_hmap = None
orig_headers = []
all_rows = []       # list of (original_values_tuple, repair_tuple)

# Counters
cnt_ok            = 0
cnt_repaired      = 0
cnt_attributed    = 0
cnt_unresolvable  = 0
cnt_needs_review  = 0
cnt_pre2016       = 0
cnt_total         = 0

REPORT_EVERY = 5000

for row in ws_baby.iter_rows(values_only=True):
    if baby_hmap is None:
        baby_hmap = header_map(row)
        orig_headers = list(row)
        print(f"  Columns ({len(orig_headers)}): {[str(h) for h in orig_headers]}")
        # Resolve column indices
        jc_idx      = get_col(baby_hmap, "job code", "job code ")
        cust_idx    = get_col(baby_hmap, "invoiced custumer", "invoiced customer", "customer")
        state_idx   = get_col(baby_hmap, "shippings state", "shipping state", "state")
        city_idx    = get_col(baby_hmap, "shipping city", "city")
        date_idx    = get_col(baby_hmap, "date shipped", "date")
        print(f"  Key cols: job_code={jc_idx}, customer={cust_idx}, state={state_idx}, "
              f"city={city_idx}, date={date_idx}")
        continue

    cnt_total += 1
    if cnt_total % REPORT_EVERY == 0:
        print(f"  ...{cnt_total:,} rows processed")

    # Extract key fields
    raw_code  = safe(row[jc_idx]    if jc_idx    is not None else None)
    row_cust  = safe(row[cust_idx]  if cust_idx  is not None else None)
    row_state = safe(row[state_idx] if state_idx is not None else None).upper()
    row_city  = safe(row[city_idx]  if city_idx  is not None else None).lower()
    row_date  = to_date(row[date_idx] if date_idx is not None else None)

    norm_code = normalize_code(raw_code)

    if norm_code and norm_code in valid_codes:
        # Sub-task A: already valid
        repair = (norm_code, "OK", "n/a", "", "")
        cnt_ok += 1

    elif raw_code == "":
        # Sub-task B: null job code
        # Pre-2016: job codes didn't exist yet — don't score, just flag
        if row_date and row_date < PRE_JOBCODE_CUTOFF:
            repair = ("", "PRE-2016", "n/a", "",
                      f"Shipped {row_date} — predates job code system (pre-2016)")
            cnt_pre2016 += 1
        elif row_date is None:
            # Unknown date with null code — attempt scoring but note missing date
            best_code, score, method, notes = find_best_match(
                "", row_cust, row_state, row_city, row_date, is_null=True
            )
            if score >= 2:
                repair = (best_code, "ATTRIBUTED", confidence(score), method,
                          f"[No ship date] {notes}")
                cnt_attributed += 1
            else:
                repair = ("", "UNRESOLVABLE", "n/a", method,
                          notes or "Null code, no date, no match found")
                cnt_unresolvable += 1
        else:
            best_code, score, method, notes = find_best_match(
                "", row_cust, row_state, row_city, row_date, is_null=True
            )
            if score >= 2:
                repair = (best_code, "ATTRIBUTED", confidence(score), method, notes)
                cnt_attributed += 1
            else:
                repair = ("", "UNRESOLVABLE", "n/a", method,
                          notes or "Null code, no match found")
                cnt_unresolvable += 1

    else:
        # Check manual review override first
        raw_upper = raw_code.strip().upper()
        # Pre-2016 partial codes: don't try to repair, just flag
        if row_date and row_date < PRE_JOBCODE_CUTOFF:
            repair = ("", "PRE-2016", "n/a", "",
                      f"Partial code '{raw_code}' on pre-2016 row ({row_date}) — predates job code system")
            cnt_pre2016 += 1
            all_rows.append((row, repair))
            continue

        if raw_upper in MANUAL_REVIEW_PARTIALS:
            # Build candidate list for notes but don't auto-assign
            candidates = set(prefix_index.get(raw_upper, []))
            cand_str = ", ".join(sorted(candidates)) if candidates else "none found"
            repair = ("", "NEEDS_REVIEW", "Manual",
                      "Prefix", f"'{raw_code}' flagged for manual review. Candidates: {cand_str}")
            cnt_needs_review += 1
            all_rows.append((row, repair))
            continue

        # Sub-task A: non-null but invalid/partial code
        best_code, score, method, notes = find_best_match(
            raw_code, row_cust, row_state, row_city, row_date, is_null=False
        )
        # Also handle case where raw_code is already valid but just wrong case
        if not best_code and norm_code and norm_code not in valid_codes:
            notes = f"'{raw_code}' not in registry, no prefix match found"

        if score >= 2:
            repair = (best_code, "REPAIRED", confidence(score), method,
                      f"'{raw_code}' -> '{best_code}'. {notes}")
            cnt_repaired += 1
        elif score == 1:
            repair = (best_code, "REPAIRED", "Low", method,
                      f"'{raw_code}' -> '{best_code}' (low confidence). {notes}")
            cnt_repaired += 1
        else:
            repair = ("", "UNRESOLVABLE", "n/a", method,
                      f"'{raw_code}' not fixable. {notes}")
            cnt_unresolvable += 1

    all_rows.append((row, repair))

wb_baby.close()
print(f"  Done. {cnt_total:,} rows processed.")
print(f"  OK:            {cnt_ok:,}")
print(f"  REPAIRED:      {cnt_repaired:,}")
print(f"  ATTRIBUTED:    {cnt_attributed:,}")
print(f"  NEEDS_REVIEW:  {cnt_needs_review:,}")
print(f"  PRE-2016:      {cnt_pre2016:,}")
print(f"  UNRESOLVABLE:  {cnt_unresolvable:,}")


# ── 4. Write output Excel ─────────────────────────────────────────────────────

print()
print("[3/3] Writing output files...")

# Style constants
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
NEW_FILL      = PatternFill("solid", fgColor="FFF2CC")   # yellow — new columns
NEW_HDR_FILL  = PatternFill("solid", fgColor="7030A0")   # purple — new column headers
ALT_FILL      = PatternFill("solid", fgColor="EEF2F7")
OK_FILL       = PatternFill("solid", fgColor="E2EFDA")   # light green
REP_FILL      = PatternFill("solid", fgColor="FCE4D6")   # light orange
ATTR_FILL     = PatternFill("solid", fgColor="DDEBF7")   # light blue
UNRES_FILL    = PatternFill("solid", fgColor="F4CCCC")   # light red
REVIEW_FILL   = PatternFill("solid", fgColor="FFD966")   # amber — needs review
PRE16_FILL    = PatternFill("solid", fgColor="D9D9D9")   # gray — pre-2016 era
thin          = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)
ALIGN_CTR     = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center")

STATUS_FILL = {
    "OK":           OK_FILL,
    "REPAIRED":     REP_FILL,
    "ATTRIBUTED":   ATTR_FILL,
    "UNRESOLVABLE": UNRES_FILL,
    "NEEDS_REVIEW": REVIEW_FILL,
    "PRE-2016":     PRE16_FILL,
}

NEW_COLS = ["Fixed_Job_Code", "Repair_Status", "Repair_Confidence", "Repair_Method", "Repair_Notes"]

wb_out = Workbook(write_only=True)
ws_out = wb_out.create_sheet("Annotated Master List")

from openpyxl.cell.cell import WriteOnlyCell

def make_cell(ws, value, font=None, fill=None, alignment=None, border=None):
    c = WriteOnlyCell(ws, value=value)
    if font:      c.font      = font
    if fill:      c.fill      = fill
    if alignment: c.alignment = alignment
    if border:    c.border    = border
    return c

# Header row
header_cells = []
for h in orig_headers:
    header_cells.append(make_cell(ws_out, h, font=HEADER_FONT, fill=HEADER_FILL,
                                  alignment=ALIGN_CTR, border=BORDER))
for nc in NEW_COLS:
    header_cells.append(make_cell(ws_out, nc, font=HEADER_FONT, fill=NEW_HDR_FILL,
                                  alignment=ALIGN_CTR, border=BORDER))
ws_out.append(header_cells)

# Data rows
for row_idx, (orig_row, repair) in enumerate(all_rows, start=2):
    status = repair[1]
    row_fill = STATUS_FILL.get(status, None) if status != "OK" else None

    cells = []
    for val in orig_row:
        c = WriteOnlyCell(ws_out, value=val)
        c.font      = BODY_FONT
        c.alignment = ALIGN_LEFT
        c.border    = BORDER
        if row_fill:
            c.fill = row_fill
        cells.append(c)

    for i, val in enumerate(repair):
        c = WriteOnlyCell(ws_out, value=val)
        c.font      = BODY_FONT
        c.alignment = ALIGN_CTR if i <= 2 else ALIGN_LEFT
        c.border    = BORDER
        if row_fill:
            c.fill = row_fill
        elif status == "OK":
            c.fill = ALT_FILL if (row_idx % 2 == 0) else NEW_FILL
        else:
            c.fill = row_fill
        cells.append(c)

    ws_out.append(cells)

wb_out.save(OUTPUT_PATH)
print(f"  Saved: {OUTPUT_PATH}")


# ── 5. Write repair log CSV ───────────────────────────────────────────────────

log_rows_written = 0
with open(LOG_PATH, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow([
        "row_num", "original_job_code", "date_shipped", "invoiced_customer",
        "state", "city", "fixed_job_code", "repair_status",
        "repair_confidence", "repair_method", "repair_notes"
    ])
    for row_idx, (orig_row, repair) in enumerate(all_rows, start=2):
        status = repair[1]
        if status == "OK":
            continue
        jc_val   = safe(orig_row[jc_idx]    if jc_idx    is not None else None)
        date_val = safe(orig_row[date_idx]   if date_idx  is not None else None)
        cust_val = safe(orig_row[cust_idx]   if cust_idx  is not None else None)
        st_val   = safe(orig_row[state_idx]  if state_idx is not None else None)
        ci_val   = safe(orig_row[city_idx]   if city_idx  is not None else None)
        writer.writerow([row_idx, jc_val, date_val, cust_val, st_val, ci_val,
                         repair[0], repair[1], repair[2], repair[3], repair[4]])
        log_rows_written += 1

print(f"  Saved: {LOG_PATH}  ({log_rows_written:,} non-OK rows logged)")


# ── Summary ───────────────────────────────────────────────────────────────────

print()
print("=" * 60)
print("PHASE 2 SUMMARY")
print("=" * 60)
print(f"Total rows processed:         {cnt_total:,}")
print(f"OK (valid code):              {cnt_ok:,}  ({100*cnt_ok/cnt_total:.1f}%)")
print(f"REPAIRED (partial code):      {cnt_repaired:,}")
print(f"ATTRIBUTED (null, 2016+):     {cnt_attributed:,}")
print(f"NEEDS_REVIEW (manual flags):  {cnt_needs_review:,}")
print(f"PRE-2016 (no job code era):   {cnt_pre2016:,}")
print(f"UNRESOLVABLE:                 {cnt_unresolvable:,}")
print()
print("Repair confidence breakdown (non-OK rows):")

conf_counts = defaultdict(int)
for _, repair in all_rows:
    if repair[1] != "OK":
        conf_counts[repair[2]] += 1
for conf, cnt in sorted(conf_counts.items()):
    print(f"  {conf:<12}: {cnt:,}")

print()
print(f"Output Excel: {OUTPUT_PATH}")
print(f"Repair log:   {LOG_PATH}")
print()
print("!! PAUSE & REVIEW CHECKPOINT !!")
print("Review jobcode_repair_log.csv before proceeding to Phase 3.")
print("Validate REPAIRED proposals and decide on UNRESOLVABLE handling.")
print("Done.")
