"""
Build a single unioned XLSX workbook from both BOM extraction projects.

Source split (no overlap):
  M Drive          2016-2022  Extracting-M-Drive/output/all_bom_{year}.csv
  BOM Str. Detail  2023-2026  BOM-Structure-Detail/output/bom_manhole_map_{year}.csv
                              (2026 has no year suffix: bom_manhole_map.csv)

Output: output\all_bom_union.xlsx

All values written as explicit strings so Excel cannot auto-convert
date-like product numbers (e.g. "12-08") to dates.
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

MDRIVE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
BSD_DIR    = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "BOM-Structure-Detail", "output")
)
OUTPUT_DIR = MDRIVE_DIR
OUT_FILE   = os.path.join(OUTPUT_DIR, "all_bom_union.xlsx")

SOURCES = [
    (MDRIVE_DIR, range(2016, 2023), False),   # M Drive: 2016-2022
    (BSD_DIR,    range(2023, 2027), True),    # BOM Structure Detail: 2023-2026
]

# M Drive has a "Zip Code" column after "Location Source" that BSD lacks.
# Inserting a blank at this position normalises all rows to the M Drive schema.
BSD_ZIP_INSERT_POS = 8  # 0-based index where "Zip Code" sits in the M Drive header


def _filename_for(directory, year, is_bsd):
    if is_bsd:
        suffix = "" if str(year) == "2026" else f"_{year}"
        return os.path.join(directory, f"bom_manhole_map{suffix}.csv")
    return os.path.join(directory, f"all_bom_{year}.csv")


def _iter_rows(directory, years, is_bsd):
    for year in years:
        path = _filename_for(directory, year, is_bsd)
        if not os.path.exists(path):
            print(f"  {year}: not found ({os.path.basename(path)}) — skipping")
            continue
        with open(path, encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            count = 0
            for row in reader:
                if is_bsd:
                    row = row[:BSD_ZIP_INSERT_POS] + [""] + row[BSD_ZIP_INSERT_POS:]
                yield row
                count += 1
        print(f"  {year}: {count:,} rows")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("BOM 2016-2026")

    # Write header from first available file
    header_written = False
    for directory, years, is_bsd in SOURCES:
        if header_written:
            break
        for year in years:
            path = _filename_for(directory, year, is_bsd)
            if os.path.exists(path):
                with open(path, encoding="utf-8", newline="") as f:
                    header = next(csv.reader(f))
                ws.append(header)
                header_written = True
                break

    total = 0
    for label, (directory, years, is_bsd) in zip(("M Drive 2016-2022", "BOM Str. Detail 2023-2026"), SOURCES):
        print(f"\n{label}:")
        for row in _iter_rows(directory, years, is_bsd):
            ws.append(row)
            total += 1

    wb.save(OUT_FILE)
    print(f"\nWrote {total:,} rows -> {OUT_FILE}")


if __name__ == "__main__":
    main()
