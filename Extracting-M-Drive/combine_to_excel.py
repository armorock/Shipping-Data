"""
Combine per-year BOM CSVs into a single XLSX workbook.

All CSV values are written as explicit strings so Excel cannot
auto-convert date-like product numbers (e.g. 12-08) to dates.
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "output"
YEARS = range(2016, 2025)
OUT_XLSX = os.path.join(OUTPUT_DIR, "all_bom_2016_2024.xlsx")


def main():
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("BOM 2016-2024")

    header_written = False
    total = 0

    for year in YEARS:
        path = os.path.join(OUTPUT_DIR, f"all_bom_{year}.csv")
        if not os.path.exists(path):
            print(f"  {year}: not found — skipping")
            continue
        with open(path, encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            header = next(reader)
            if not header_written:
                ws.append(header)
                header_written = True
            for row in reader:
                ws.append(row)
                total += 1
        print(f"  {year}: loaded")

    wb.save(OUT_XLSX)
    print(f"\nWrote {total:,} rows -> {OUT_XLSX}")


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
