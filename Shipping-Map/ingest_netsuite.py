import csv
import glob
import os
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import date, datetime

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
DB_DATA = os.path.normpath(os.path.join(_HERE, "..", "Combined-Database", "data"))
XLSX_PATH = r"C:\Users\JohnLeitzke\OneDrive - Armorock LLC\Documents\Desktop\NSAW All Shipping Data1.2.xlsx"
GEONAMES_PATH = os.path.join(_HERE, "data", "US.txt")

PRODUCT_TYPE_MAP = {
    "Base":        "MHB",
    "Riser":       "MHS",
    "Short Riser": "MHS",
    "Lid":         "MHL",
    "Short Lid":   "MHL",
    "Cone":        "MHC",
    "Rehab Riser": "RMH",
}

PLANTS = {"Boulder City", "Sulphur Springs", "Plant City"}
SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"


def _tag(local):
    return f"{{{SS_NS}}}{local}"


def _attr(el, local):
    v = el.get(f"{{{SS_NS}}}{local}")
    return v if v is not None else el.get(local)


def parse_xls(path):
    tree = ET.parse(path)
    root = tree.getroot()
    ws = root.find(f".//{_tag('Worksheet')}")
    if ws is None:
        sys.exit(f"ERROR: no Worksheet element found in {path}")
    table = ws.find(f".//{_tag('Table')}")
    if table is None:
        sys.exit(f"ERROR: no Table element found in {path}")
    header = None
    records = []
    for row_el in table.findall(_tag("Row")):
        cells = []
        col = 0
        for cell_el in row_el.findall(_tag("Cell")):
            idx_s = _attr(cell_el, "Index")
            if idx_s:
                target = int(idx_s) - 1
                while len(cells) < target:
                    cells.append("")
                col = target
            data_el = cell_el.find(_tag("Data"))
            cells.append(data_el.text if data_el is not None and data_el.text else "")
            col += 1
        if header is None:
            header = [c.strip() for c in cells]
            continue
        row = {h: (cells[i] if i < len(cells) else "") for i, h in enumerate(header)}
        records.append(row)
    return records


def parse_date(s):
    if not s:
        return None
    s = s.split("T")[0]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def norm_state(v):
    if not v:
        return None
    s = str(v).strip().upper()
    return s if len(s) == 2 else None


def norm_plant(v):
    if not v:
        return None
    s = str(v).strip()
    for p in PLANTS:
        if s.upper().startswith(p.upper()):
            return p
    return s


def load_geonames_city_states():
    city_states = defaultdict(set)
    with open(GEONAMES_PATH, encoding="utf-8") as f:
        for row in csv.reader(f, delimiter="\t"):
            city, state = row[2].upper().strip(), row[4].strip()
            if city and state and len(state) == 2:
                city_states[city].add(state)
    return city_states


def nsaw_cutoff():
    if not os.path.exists(XLSX_PATH):
        sys.exit(f"ERROR: NSAW xlsx not found:\n  {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    header = None
    cutoff = None
    for raw in ws.iter_rows(values_only=True):
        if header is None:
            header = {str(h).strip(): i for i, h in enumerate(raw) if h is not None}
            continue
        d = raw[header["Date Shipped"]]
        if isinstance(d, datetime) and (cutoff is None or d > cutoff):
            cutoff = d
        elif isinstance(d, date) and not isinstance(d, datetime):
            dt = datetime(d.year, d.month, d.day)
            if cutoff is None or dt > cutoff:
                cutoff = dt
    wb.close()
    return cutoff


def main():
    os.makedirs(os.path.join(_HERE, "output"), exist_ok=True)

    appx_path = (
        find_latest(os.path.join(DB_DATA, "AppxShippedProduct*.xls"))
        or find_latest(os.path.join(DB_DATA, "AppxShipped*.xls"))
    )
    os_path = find_latest(os.path.join(DB_DATA, "OSLocationData*.xls"))

    if not appx_path or not os_path:
        sys.exit(
            "ERROR: XLS files not found in Combined-Database/data/\n"
            "  Expected: AppxShipped*.xls and OSLocationData*.xls"
        )

    print(f"AppxShipped : {os.path.basename(appx_path)}")
    print(f"OSLocation  : {os.path.basename(os_path)}")

    appx_rows = parse_xls(appx_path)
    os_rows   = parse_xls(os_path)
    print(f"Parsed {len(appx_rows):,} AppxShipped rows, {len(os_rows):,} OS rows")

    os_by_sit = {}
    for r in os_rows:
        sit = r.get("Shipment Item Transaction", "").strip()
        if sit:
            os_by_sit[sit] = r

    joined, no_match = [], 0
    for r in appx_rows:
        sit = r.get("Shipment Item Transaction", "").strip()
        os_r = os_by_sit.get(sit)
        if os_r is None:
            no_match += 1
        else:
            joined.append((r, os_r))
    print(f"Joined {len(joined):,} rows ({no_match} had no OS match)")

    cutoff = nsaw_cutoff()
    print(f"NSAW cutoff : {cutoff.date() if cutoff else 'none'}")

    new_rows = []
    skipped_date, skipped_cutoff = 0, 0
    for appx, os_r in joined:
        d = parse_date(appx.get("Date", ""))
        if d is None:
            skipped_date += 1
            continue
        if cutoff and d <= cutoff:
            skipped_cutoff += 1
            continue
        new_rows.append((appx, os_r, d))

    print(f"After cutoff filter: {len(new_rows):,} new records "
          f"(skipped {skipped_cutoff:,} before/at cutoff, {skipped_date} bad dates)")

    if not new_rows:
        print("Nothing to append.")
        return

    city_states = load_geonames_city_states()
    anomalies = []

    final_rows = []
    for appx, os_r, d in new_rows:
        pt_raw = appx.get("Product Type", "").strip()
        pt = PRODUCT_TYPE_MAP.get(pt_raw, pt_raw)

        city = os_r.get("City", "").strip() or None
        state = norm_state(os_r.get("State", ""))
        zip_val = os_r.get("Zip Code", "").strip() or None

        if city and state:
            valid = city_states.get(city.upper(), set())
            if valid and state not in valid:
                anomalies.append({
                    "job_code":     appx.get("Job Code", "").strip(),
                    "city":         city,
                    "listed_state": state,
                    "valid_states": "|".join(sorted(valid)),
                    "zip":          zip_val or "",
                    "date":         d.date().isoformat(),
                    "part_type":    pt,
                })

        qty_s = os_r.get("Quantity Fufilled", "").strip()
        try:
            qty = int(float(qty_s)) if qty_s else 1
        except ValueError:
            qty = 1

        final_rows.append({
            "Quantity":          qty,
            "Date Shipped":      d,
            "Plant":             norm_plant(appx.get("Location", "")),
            "diameter":          appx.get("Product Diameter", "").strip(),
            "Invoiced Custumer": appx.get("Name", "").strip(),
            "Job Code":          appx.get("Job Code", "").strip(),
            "Structure Name":    appx.get("Structure Type", "").strip(),
            "Part Number":       appx.get("Item", "").strip(),
            "part_type":         pt,
            "Shipping City":     city or "",
            "Shippings State":   state or "",
            "ZipCode":           zip_val or "",
            "Current Part Name": appx.get("Item", "").strip(),
        })

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Sheet1"]
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_order = [str(h).strip() for h in header_row if h is not None]

    for r in final_rows:
        ws.append([r.get(col, "") for col in col_order])

    wb.save(XLSX_PATH)

    anom_path = os.path.join(_HERE, "output", "geo_anomalies.csv")
    if anomalies:
        with open(anom_path, "w", newline="", encoding="utf-8") as f:
            fields = ["job_code", "city", "listed_state", "valid_states", "zip", "date", "part_type"]
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(anomalies)

    by_type = Counter(r["part_type"] for r in final_rows)
    dates   = [r["Date Shipped"].date() for r in final_rows]
    print(f"\nAppended {len(final_rows):,} rows to NSAW xlsx "
          f"({min(dates)} to {max(dates)})")
    for pt, n in by_type.most_common():
        print(f"  {pt}: {n:,}")
    if anomalies:
        print(f"\nWARN: {len(anomalies)} suspect city/state combos -> {anom_path}")
    else:
        print("\nNo geo anomalies in new records.")


def find_latest(pattern):
    files = glob.glob(pattern)
    return max(files, key=os.path.getmtime) if files else None


if __name__ == "__main__":
    main()
