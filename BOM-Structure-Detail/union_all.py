"""
union_all.py — combine per-year BOM CSVs into a single XLSM workbook.

Usage:
    python union_all.py                   # unions 2023–2026
    python union_all.py 2025 2026         # specific years only
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "output"
YEARS = ["2023", "2024", "2025", "2026"]
OUT_FILE = os.path.join(OUTPUT_DIR, "all_bom_2023_2026.xlsm")

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _col_widths(df, max_sample=500):
    widths = {}
    for col in df.columns:
        header_len = len(str(col))
        sample = df[col].dropna().astype(str)
        if len(sample) > max_sample:
            sample = sample.sample(max_sample, random_state=0)
        data_len = sample.str.len().max() if len(sample) else 0
        widths[col] = min(max(header_len, int(data_len or 0)) + 2, 60)
    return widths


def main():
    years = sys.argv[1:] if len(sys.argv) > 1 else YEARS

    dfs = []
    for year in years:
        path = os.path.join(OUTPUT_DIR, f"all_bom_{year}.csv")
        if not os.path.exists(path):
            print(f"  {year}: not found — skipping ({path})")
            continue
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
        print(f"  {year}: {len(df):,} rows")
        dfs.append(df)

    if not dfs:
        print("No data found.")
        sys.exit(1)

    combined = pd.concat(dfs, ignore_index=True)
    print(f"\nTotal: {len(combined):,} rows  x  {len(combined.columns)} columns")

    widths = _col_widths(combined)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Writing workbook…")
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "BOM Data"

    headers = list(combined.columns)
    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for col_idx, col_name in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_name]

    for row in combined.itertuples(index=False, name=None):
        ws.append(list(row))

    wb.save(OUT_FILE)
    print(f"Saved: {OUT_FILE}")
    print("Open in Excel and Ctrl+S to finalize as XLSM if you need macro support.")


if __name__ == "__main__":
    main()
