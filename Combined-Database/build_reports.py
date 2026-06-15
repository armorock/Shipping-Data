import csv
import os
import sqlite3
from collections import Counter, defaultdict

import openpyxl

import sources
import resolutions
from common import norm_job_code, year_from_date, structure_class, part_type_from_pn, is_blank

LEDGER = os.path.join(sources.OUTPUT_DIR, "pieces_ledger.csv")
REGISTRY = os.path.join(sources.OUTPUT_DIR, "project_registry.csv")
CLASSES = ["base", "rehab", "non-base", "unknown"]
REVIEW_SOURCES = ["bom_union", "dispatch", "erp_qb", "erp_fb", "erp_ns", "jobcode_db", "registry"]


def _load_geonames():
    gz = {}
    try:
        with open(sources.GEONAMES_TXT, encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split("\t")
                if len(parts) >= 5:
                    county = parts[5].upper() if len(parts) > 5 and parts[5] else ""
                    gz.setdefault(parts[1], (parts[2].upper(), parts[4].upper(), county))
    except FileNotFoundError:
        pass
    return gz


def _ledger():
    return list(csv.DictReader(open(LEDGER, encoding="utf-8")))


def _freeze(ws):
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------- rehab / non-base counts
def rehab_nonbase_report(rows):
    by_year = defaultdict(Counter)
    by_state = defaultdict(Counter)
    by_cust = defaultdict(Counter)
    reg_cust = {r["job_code"]: r.get("customer") for r in csv.DictReader(open(REGISTRY, encoding="utf-8"))}
    for p in rows:
        cls = p["structure_class"]
        by_year[p["year"] or "?"][cls] += 1
        by_state[p["state"] or "?"][cls] += 1
        by_cust[reg_cust.get(p["job_code"]) or "?"][cls] += 1

    wb = openpyxl.Workbook()

    def sheet(title, data, key_name, sort_total=False):
        ws = wb.create_sheet(title)
        ws.append([key_name] + CLASSES + ["total"])
        keys = sorted(data, key=lambda k: -sum(data[k].values())) if sort_total else sorted(data, key=str)
        for k in keys:
            c = data[k]
            ws.append([k] + [c[cl] for cl in CLASSES] + [sum(c.values())])
        _freeze(ws)

    wb.remove(wb.active)
    sheet("by_year", by_year, "year")
    sheet("by_state", by_state, "state", sort_total=True)
    sheet("by_customer", by_cust, "customer", sort_total=True)
    total = Counter(p["structure_class"] for p in rows)
    ws = wb.create_sheet("totals", 0)
    ws.append(["class", "pieces"])
    for cl in CLASSES:
        ws.append([cl, total[cl]])
    ws.append(["TOTAL", sum(total.values())])
    path = os.path.join(sources.OUTPUT_DIR, "rehab_nonbase_report.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------- provenance / needs-review
def project_provenance():
    con = sqlite3.connect(sources.OBSERVATIONS_DB)
    conf = {}
    gap = Counter()
    for ek, field, status in con.execute("SELECT entity_key, field, status FROM entity_field_resolution"):
        if status == "CONFLICT":
            conf[ek] = conf.get(ek, 0) + 1
    con.close()
    reg = list(csv.DictReader(open(REGISTRY, encoding="utf-8")))
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "provenance"
    ws.append(["job_code", "sources", "n_sources", "location_conflicts", "city", "state",
               "city_status", "state_status", "needs_review"])
    for r in reg:
        srcs = r.get("sources") or ""
        ws.append([r["job_code"], srcs, len(srcs.split(",")) if srcs else 0,
                   conf.get(r["job_code"], 0), r.get("city"), r.get("state"),
                   r.get("city_status"), r.get("state_status"), r.get("needs_review")])
    _freeze(ws)
    path = os.path.join(sources.OUTPUT_DIR, "project_provenance.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------- structure completeness roadmap
DOC_DEPTH = {  # Source File Name -> depth tier rank (higher = richer)
    "Shop Drawing PDF": ("DIAGRAM", 4),
    "BOM by Structure XML": ("PER_STRUCTURE_BOM", 3),
    "BOM by Structure PDF": ("PER_STRUCTURE_BOM", 3),
    "BOM Summary PDF": ("SUMMARY_ONLY", 2),
    "Quotation PDF": ("SUMMARY_ONLY", 2),
}


def structure_completeness_roadmap(rows):
    # shipped structures per job (from ledger structure_id)
    shipped_struct = defaultdict(Counter)   # job -> Counter(structure_class) for pieces w/ structure
    shipped_any = defaultdict(Counter)      # job -> Counter(structure_class) all pieces
    for p in rows:
        shipped_any[p["job_code"]][p["structure_class"]] += 1

    # BOM structures + best doc depth per (job, structure)
    wb = openpyxl.load_workbook(sources.BOM_UNION_XLSX, read_only=True, data_only=True)
    ws = wb[sources.BOM_UNION_SHEET]
    it = ws.iter_rows(values_only=True)
    h = {str(c).strip(): i for i, c in enumerate(next(it)) if c is not None}
    struct = {}   # (job, structure) -> {depth, parts:Counter(class)}
    for row in it:
        jc = norm_job_code(row[h["Job Code"]])
        sname = row[h["Structure Name"]]
        if not jc or is_blank(sname):
            continue
        key = (jc, str(sname).strip())
        tier, rank = DOC_DEPTH.get(str(row[h["Source File Name"]]).strip(), ("SUMMARY_ONLY", 1))
        rec = struct.setdefault(key, {"tier": tier, "rank": rank, "parts": Counter()})
        if rank > rec["rank"]:
            rec["tier"], rec["rank"] = tier, rank
        cls = structure_class(row[h["Part Type"]], row[h["Product Number"]])
        rec["parts"][cls] += 1
    wb.close()

    out = openpyxl.Workbook(); wsx = out.active; wsx.title = "roadmap"
    wsx.append(["job_code", "structure", "data_depth", "has_base", "has_section_nonbase",
                "bom_part_count", "shipped_in_job", "next_action"])
    shipped_jobs = set(shipped_any)
    tier_counts = Counter()
    for (jc, sname), rec in sorted(struct.items()):
        has_base = rec["parts"]["base"] > 0
        has_nb = rec["parts"]["non-base"] > 0
        tier_counts[rec["tier"]] += 1
        if rec["tier"] in ("DIAGRAM", "PER_STRUCTURE_BOM"):
            action = "complete detail on file"
            if not has_base and rec["parts"]["rehab"] == 0:
                action = "no base in BOM - verify (rehab or missing base?)"
        elif rec["tier"] == "SUMMARY_ONLY":
            action = "summary only - pull BOM by Structure / Shop Drawing from job release folder"
        else:
            action = "no per-structure detail - pull Shop Drawing from SharePoint/M-drive"
        wsx.append([jc, sname, rec["tier"], "Y" if has_base else "", "Y" if has_nb else "",
                    sum(rec["parts"].values()), "Y" if jc in shipped_jobs else "", action])

    # Jobs whose shipped pieces have NO per-structure BOM at all (SHIPPING_ONLY).
    bom_jobs = {jc for (jc, _) in struct}
    for jc in sorted(shipped_jobs - bom_jobs):
        tier_counts["SHIPPING_ONLY"] += 1
        wsx.append([jc, "(no structure detail)", "SHIPPING_ONLY", "", "",
                    0, "Y", "shipped but no BOM/diagram found - pull BOM by Structure / Shop Drawing"])
    _freeze(wsx)
    path = os.path.join(sources.OUTPUT_DIR, "structure_completeness_roadmap.xlsx")
    out.save(path)
    return path, tier_counts


# ---------------------------------------------------------------- per-scope review lists
def review_lists():
    con = sqlite3.connect(sources.OBSERVATIONS_DB)
    confirmed = resolutions.load()
    counts = {}
    geonames = _load_geonames()

    project_names = {row[0]: row[1] for row in con.execute(
        "SELECT entity_key, winner FROM entity_field_resolution WHERE field='project_name'"
    )}

    def src_values(ek, field):
        rows = con.execute(
            "SELECT source, value_norm FROM observations WHERE entity_key=? AND field=?"
            " ORDER BY trust DESC, n_pieces DESC",
            (ek, field),
        ).fetchall()
        seen = {}
        for src, val in rows:
            if src not in seen:
                seen[src] = val
        return [seen.get(s, "") for s in REVIEW_SOURCES]

    def export(field, fname, statuses, extra_headers=None, extra_row_fn=None):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = field
        src_headers = [f"src_{s}" for s in REVIEW_SOURCES]
        ws.append(
            ["job_code", "field", "status", "current_winner", "winner_sources", "project_name"]
            + (extra_headers or [])
            + src_headers
            + ["all_observed_values", "Confirmed Value", "Notes"]
        )
        q = con.execute(
            "SELECT entity_key, status, winner, winner_sources"
            " FROM entity_field_resolution WHERE field=? ORDER BY status DESC, entity_key",
            (field,),
        )
        n = 0
        for ek, status, winner, wsrcs in q:
            if status not in statuses or (ek, field) in confirmed:
                continue
            obs = con.execute(
                "SELECT value_norm, source, trust, n_pieces FROM observations"
                " WHERE entity_key=? AND field=? ORDER BY trust DESC",
                (ek, field),
            ).fetchall()
            allv = " | ".join(f"{v} [{s}·t{t:.2f}·n{n2}]" for v, s, t, n2 in obs) or "(none)"
            extra = extra_row_fn(ek, winner) if extra_row_fn else []
            ws.append(
                [ek, field, status, winner, wsrcs, project_names.get(ek, "")]
                + extra
                + src_values(ek, field)
                + [allv, "", ""]
            )
            n += 1
        _freeze(ws)
        wb.save(os.path.join(sources.REVIEW_DIR, fname))
        counts[fname] = n

    def city_extras(ek, winner):
        bom_row = con.execute(
            "SELECT value_raw FROM observations"
            " WHERE entity_key=? AND field='city' AND source='bom_union' LIMIT 1",
            (ek,),
        ).fetchone()
        raw_bom = bom_row[0] if bom_row else ""
        zip_row = con.execute(
            "SELECT winner FROM entity_field_resolution WHERE entity_key=? AND field='zip'", (ek,)
        ).fetchone()
        confirmed_zip = zip_row[0] if zip_row else None
        _, _, zip_county = geonames.get(str(confirmed_zip), ("", "", "")) if confirmed_zip else ("", "", "")
        county_flag = winner if (winner and "COUNTY" in str(winner).upper().split()) else ""
        return [raw_bom, zip_county, county_flag]

    def zip_extras(ek, winner):
        geo_city, geo_state, _ = geonames.get(str(winner), ("", "", "")) if winner else ("", "", "")
        city_row = con.execute(
            "SELECT winner FROM entity_field_resolution WHERE entity_key=? AND field='city'", (ek,)
        ).fetchone()
        state_row = con.execute(
            "SELECT winner FROM entity_field_resolution WHERE entity_key=? AND field='state'", (ek,)
        ).fetchone()
        conf_city = city_row[0] if city_row else ""
        conf_state = state_row[0] if state_row else ""
        if not geo_city:
            zip_ok = "UNKNOWN_ZIP"
        elif geo_state != conf_state:
            zip_ok = "STATE_MISMATCH"
        elif conf_city and geo_city != conf_city:
            zip_ok = "CITY_MISMATCH"
        else:
            zip_ok = "✓"
        return [geo_city, geo_state, conf_city, conf_state, zip_ok]

    def location_matrix():
        ek_set = {row[0] for row in con.execute(
            "SELECT DISTINCT entity_key FROM observations"
            " WHERE field IN ('city','state','zip','county')"
        )}
        res = {}
        for ek, field, status, winner in con.execute(
            "SELECT entity_key, field, status, winner FROM entity_field_resolution"
            " WHERE field IN ('city','state','zip','county','project_name')"
        ):
            res.setdefault(ek, {})[field] = (status, winner)
        src_vals = {}
        for ek, field, source, val in con.execute(
            "SELECT entity_key, field, source, value_norm FROM observations"
            " WHERE field IN ('city','state','county')"
            " ORDER BY trust DESC, n_pieces DESC"
        ):
            key = (ek, field, source)
            if key not in src_vals:
                src_vals[key] = val
        street_addrs = {}
        for ek, val in con.execute(
            "SELECT entity_key, value_norm FROM observations WHERE field='street_address'"
            " ORDER BY trust DESC"
        ):
            street_addrs.setdefault(ek, val)

        SOURCES = ["bom_union", "dispatch", "erp_qb", "erp_fb", "erp_ns", "jobcode_db", "registry"]
        LOC = ["city", "state", "county"]
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "location_matrix"
        header = ["job_code", "project_name", "city_status", "state_status", "county_status",
                  "confirmed_city", "confirmed_state", "confirmed_zip", "confirmed_county",
                  "zip_city", "zip_county", "street_address"]
        for src in SOURCES:
            for fld in LOC:
                header.append(f"{src}_{fld}")
        ws.append(header)
        for ek in sorted(ek_set):
            r = res.get(ek, {})
            city_s, city_w = r.get("city", ("GAP", None))
            state_s, state_w = r.get("state", ("GAP", None))
            county_s, county_w = r.get("county", ("GAP", None))
            _, zip_w = r.get("zip", ("GAP", None))
            _, pn_w = r.get("project_name", ("GAP", None))
            geo_city, _, geo_county = geonames.get(str(zip_w), ("", "", "")) if zip_w else ("", "", "")
            row_data = [ek, pn_w or "", city_s, state_s, county_s,
                        city_w or "", state_w or "", zip_w or "", county_w or "",
                        geo_city, geo_county, street_addrs.get(ek, "")]
            for src in SOURCES:
                for fld in LOC:
                    row_data.append(src_vals.get((ek, fld, src), "") or "")
            ws.append(row_data)
        _freeze(ws)
        path = os.path.join(sources.REVIEW_DIR, "07_location_matrix.xlsx")
        wb.save(path)
        return len(ek_set)

    export("state", "01_state.xlsx", {"CONFLICT", "GAP"})
    export("city", "02_city.xlsx", {"CONFLICT", "GAP"},
           extra_headers=["raw_bom", "zip_county", "county_flag"], extra_row_fn=city_extras)
    export("zip", "03_zip.xlsx", {"CONFLICT", "GAP"},
           extra_headers=["geo_city", "geo_state", "confirmed_city", "confirmed_state", "zip_ok"],
           extra_row_fn=zip_extras)
    export("county", "04_county.xlsx", {"CONFLICT", "GAP"})

    # 06 all conflicts across every field, ranked by number of competing values
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "conflicts_all"
    ws.append(["job_code", "field", "n_values", "current_winner", "winner_sources",
               "all_observed_values", "Confirmed Value", "Notes"])
    q = con.execute("""SELECT entity_key, field, n_values, winner, winner_sources
                       FROM entity_field_resolution WHERE status='CONFLICT'
                       AND field != 'project_name'
                       ORDER BY n_values DESC, field""")
    n06 = 0
    for ek, field, nv, winner, wsrcs in q:
        if (ek, field) in confirmed:
            continue
        obs = con.execute("""SELECT value_norm, source, trust, n_pieces FROM observations
                             WHERE entity_key=? AND field=? ORDER BY trust DESC""", (ek, field)).fetchall()
        allv = " | ".join(f"{v} [{s}·t{t:.2f}·n{n2}]" for v, s, t, n2 in obs)
        ws.append([ek, field, nv, winner, wsrcs, allv, "", ""])
        n06 += 1
    _freeze(ws)
    wb.save(os.path.join(sources.REVIEW_DIR, "06_conflicts_all.xlsx"))
    counts["06_conflicts_all.xlsx"] = n06

    counts["07_location_matrix.xlsx"] = location_matrix()
    con.close()
    return counts


def write_index(counts, tier_counts):
    lines = ["# Review index", "",
             "Each list is independent — fill the **Confirmed Value** column on the rows you want to",
             "fix, then run `apply_corrections.py <file>`. Decisions persist in `data/resolutions.csv`.",
             ""]
    for f in ("01_state.xlsx", "02_city.xlsx", "03_zip.xlsx", "04_county.xlsx", "06_conflicts_all.xlsx"):
        lines.append(f"- `{f}` — {counts.get(f, 0)} rows to review")
    lines.append(f"- `07_location_matrix.xlsx` — {counts.get('07_location_matrix.xlsx', 0)} jobs (all sources, includes AGREE rows)")
    lines += ["", "## Structure data-depth (completeness roadmap)"]
    for tier, n in sorted(tier_counts.items(), key=lambda kv: -kv[1]):
        lines.append(f"- {tier}: {n} structures")
    with open(os.path.join(sources.REVIEW_DIR, "_index.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def auto_filled_part_type(rows):
    path = os.path.join(sources.OUTPUT_DIR, "auto_filled_part_type.csv")
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["piece_id", "job_code", "part_number", "derived_part_type", "structure_class"])
        for p in rows:
            if is_blank(p.get("part_type")) and p["structure_class"] != "unknown":
                w.writerow([p["piece_id"], p["job_code"], p["part_number"],
                            part_type_from_pn(p["part_number"]), p["structure_class"]])
    return path


def build():
    sources.ensure_dirs()
    rows = _ledger()
    p1 = rehab_nonbase_report(rows)
    p2 = project_provenance()
    p3, tiers = structure_completeness_roadmap(rows)
    counts = review_lists()
    write_index(counts, tiers)
    p4 = auto_filled_part_type(rows)
    print("wrote:", os.path.basename(p1), os.path.basename(p2), os.path.basename(p3), os.path.basename(p4))
    print("review lists:", counts)
    print("structure tiers:", dict(tiers))


if __name__ == "__main__":
    build()
