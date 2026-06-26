import csv
import os
from collections import Counter, defaultdict

import openpyxl

import sources
from common import (norm_job_code, norm_part_number, norm_city, norm_state, norm_zip,
                    year_from_date, structure_class, part_type_from_pn, is_blank)
from parse_part_name import parse_part_name, build_gen4_name

csv.field_size_limit(10_000_000)


def _date_key(v):
    if v is None:
        return ""
    if hasattr(v, "date"):
        return v.date().isoformat()
    return str(v).strip()


def read_dispatch_pieces():
    """One row per physical piece (Dispatch is already piece-grain)."""
    pieces = []
    with open(sources.DISPATCH_CSV, encoding="utf-8") as f:
        for i, row in enumerate(csv.DictReader(f)):
            pn = norm_part_number(row.get("part_number"))
            if not pn:
                continue
            jc = norm_job_code(row.get("job_code"))
            year = year_from_date(row.get("year") or row.get("ship_date"))
            pt = (row.get("part_type") or "").strip().upper() or part_type_from_pn(pn)
            pieces.append({
                "source": "dispatch",
                "job_code": jc, "structure_id": (row.get("structure_id") or "").strip() or None,
                "part_number": pn, "part_type": pt,
                "structure_class": structure_class(pt, pn),
                "ship_date": _date_key(row.get("ship_date")), "year": year,
                "plant": (row.get("plant") or "").strip().upper() or None,
                "piece_id": f"dispatch:{i}",
            })
    return pieces


def read_erp_groups():
    """Group ERP rows by (job, part, date); dedup across QB/FB/NS by taking the MAX
    per-system count (same piece migrated across accounting systems is not 2 pieces)."""
    wb = openpyxl.load_workbook(sources.BABY_XLSM, read_only=True, data_only=True)
    ws = wb[sources.BABY_SHEET]
    it = ws.iter_rows(values_only=True)
    h = {str(c).strip(): i for i, c in enumerate(next(it)) if c is not None}
    groups = defaultdict(lambda: defaultdict(int))   # key -> {erp_system: qty}
    meta = {}
    for row in it:
        erp = sources.ERP_LABELS.get(str(row[h["ERP"]]).strip()) if row[h["ERP"]] else None
        if not erp:
            continue
        qty = row[h["Quantity"]]
        if qty == 0:
            continue
        pn = norm_part_number(row[h["Part Number"]])
        if not pn:                       # no part number -> not a countable piece (empty/junk row)
            continue
        q = int(qty) if isinstance(qty, (int, float)) and qty else 1   # blank qty -> 1 (NSAW rule)
        jc = norm_job_code(row[h["Job Code"]])
        date = _date_key(row[h["Date Shipped"]])
        key = (jc, pn, date)
        groups[key][erp] += q
        if key not in meta:
            pt = (str(row[h["Part Type"]]).strip().upper() if not is_blank(row[h["Part Type"]]) else None) \
                 or part_type_from_pn(pn)
            meta[key] = {
                "part_type": pt,
                "structure_id": (str(row[h["Structure Name"]]).strip()
                                 if not is_blank(row[h["Structure Name"]]) else None),
                "plant": (str(row[h["Plant"]]).strip().upper() if not is_blank(row[h["Plant"]]) else None),
                "year": year_from_date(row[h["Date Shipped"]]),
                "city":  norm_city(str(row[h["Shipping City"]]) if not is_blank(row[h["Shipping City"]]) else None),
                "state": norm_state(str(row[h["Shippings State"]]) if not is_blank(row[h["Shippings State"]]) else None),
                "zip":   norm_zip(row[h["ZipCode"]]) if not is_blank(row[h["ZipCode"]]) else None,
            }
    wb.close()
    return groups, meta


def dedup_erp_pieces(groups, meta):
    pieces = []
    for (jc, pn, date), systems in groups.items():
        n = max(systems.values())          # dedup across ERP systems
        m = meta[(jc, pn, date)]
        srcs = "+".join(sorted(systems))
        for s in range(n):
            pieces.append({
                "source": srcs, "job_code": jc, "structure_id": m["structure_id"],
                "part_number": pn, "part_type": m["part_type"],
                "structure_class": structure_class(m["part_type"], pn),
                "ship_date": date, "year": m["year"], "plant": m["plant"],
                "piece_id": f"erp:{jc}:{pn}:{date}:{s}",
                "erp_city": m["city"], "erp_state": m["state"], "erp_zip": m["zip"],
            })
    return pieces


def load_registry_loc():
    path = os.path.join(sources.OUTPUT_DIR, "project_registry.csv")
    loc = {}
    if os.path.exists(path):
        for r in csv.DictReader(open(path, encoding="utf-8")):
            loc[r["job_code"]] = (r.get("city"), r.get("state"), r.get("zip"))
    return loc


def reconciliation(dispatch, erp):
    """Compare Dispatch vs deduped-ERP piece counts for the overlap years 2019-2025."""
    yrs = sources.DISPATCH_YEARS
    d_job = Counter(); e_job = Counter()
    d_year = Counter(); e_year = Counter()
    d_jobs = defaultdict(set); e_jobs = defaultdict(set)
    for p in dispatch:
        if p["year"] in yrs:
            d_job[p["job_code"]] += 1; d_year[p["year"]] += 1; d_jobs[p["year"]].add(p["job_code"])
    for p in erp:
        if p["year"] in yrs:
            e_job[p["job_code"]] += 1; e_year[p["year"]] += 1; e_jobs[p["year"]].add(p["job_code"])

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "by_year"
    ws.append(["year", "dispatch_pieces", "erp_pieces", "diff(d-e)",
               "jobs_dispatch_only", "jobs_erp_only", "jobs_both"])
    for y in sorted(yrs):
        do = d_jobs[y] - e_jobs[y]; eo = e_jobs[y] - d_jobs[y]; both = d_jobs[y] & e_jobs[y]
        ws.append([y, d_year[y], e_year[y], d_year[y] - e_year[y], len(do), len(eo), len(both)])
    ws.append(["TOTAL", sum(d_year.values()), sum(e_year.values()),
               sum(d_year.values()) - sum(e_year.values()), "", "", ""])

    ws2 = wb.create_sheet("by_job")
    ws2.append(["job_code", "dispatch_pieces", "erp_pieces", "diff(d-e)", "present_in"])
    for jc in sorted(set(d_job) | set(e_job), key=lambda x: (x is None, x or "")):
        d, e = d_job[jc], e_job[jc]
        where = "both" if d and e else ("dispatch_only" if d else "erp_only")
        ws2.append([jc, d, e, d - e, where])
    for w in (ws, ws2):
        w.freeze_panes = "A2"
    path = os.path.join(sources.OUTPUT_DIR, "dispatch_vs_erp_reconciliation.xlsx")
    wb.save(path)
    return path, sum(d_year.values()), sum(e_year.values())


def quoted_not_shipped(shipped_jobs):
    """Jobs present in the BOM/released layer with zero shipped pieces (released/quoted, not shipped)."""
    wb = openpyxl.load_workbook(sources.BOM_UNION_XLSX, read_only=True, data_only=True)
    ws = wb[sources.BOM_UNION_SHEET]
    it = ws.iter_rows(values_only=True)
    h = {str(c).strip(): i for i, c in enumerate(next(it)) if c is not None}
    bom = {}
    for row in it:
        jc = norm_job_code(row[h["Job Code"]])
        if not jc:
            continue
        rec = bom.setdefault(jc, {"structures": set(), "project": row[h["Project Name"]],
                                  "loc": row[h["Job Location"]], "years": set(), "docs": set()})
        if not is_blank(row[h["Structure Name"]]):
            rec["structures"].add(str(row[h["Structure Name"]]).strip())
        y = year_from_date(row[h["Year Release"]])
        if y:
            rec["years"].add(y)
        if not is_blank(row[h["Source File Name"]]):
            rec["docs"].add(str(row[h["Source File Name"]]).strip())
    wb.close()

    out = openpyxl.Workbook(); wsx = out.active; wsx.title = "quoted_not_shipped"
    wsx.append(["job_code", "project_name", "job_location", "bom_years", "structures_in_bom",
                "doc_types", "note"])
    n = 0
    for jc in sorted(bom):
        if jc in shipped_jobs:
            continue
        r = bom[jc]
        quote_only = r["docs"] and r["docs"].issubset({"Quotation PDF", "Shop Drawing PDF"})
        wsx.append([jc, r["project"], r["loc"],
                    ",".join(str(y) for y in sorted(r["years"])), len(r["structures"]),
                    ",".join(sorted(r["docs"])),
                    "quote/drawing only" if quote_only else "released, no shipment"])
        n += 1
    wsx.freeze_panes = "A2"
    path = os.path.join(sources.OUTPUT_DIR, "quoted_not_shipped.xlsx")
    out.save(path)
    return path, n


def build():
    sources.ensure_dirs()
    dispatch = read_dispatch_pieces()
    groups, meta = read_erp_groups()
    erp = dedup_erp_pieces(groups, meta)
    erp_raw = sum(sum(s.values()) for s in groups.values())
    print(f"dispatch pieces: {len(dispatch)} | erp pieces raw: {erp_raw} -> deduped: {len(erp)}")

    recon_path, d_tot, e_tot = reconciliation(dispatch, erp)
    print(f"reconciliation (2019-2025): dispatch={d_tot} erp={e_tot} -> {recon_path}")

    # Apply the overlap rule.
    rule = sources.SHIPPED_OVERLAP
    erp_jobs = {p["job_code"] for p in erp if p["job_code"]}
    if rule == "dispatch":
        ledger = dispatch + [p for p in erp if p["year"] not in sources.DISPATCH_YEARS]
    elif rule == "erp":
        ledger = list(erp)
    else:  # union: ERP everywhere + pieces from jobs that appear ONLY in dispatch
        dispatch_only = [p for p in dispatch if p["job_code"] not in erp_jobs]
        ledger = erp + dispatch_only
        print(f"  union adds {len(dispatch_only)} pieces from dispatch-only jobs")
    print(f"SHIPPED_OVERLAP={rule} -> ledger pieces: {len(ledger)}")

    loc = load_registry_loc()
    for p in ledger:
        c, s, z = loc.get(p["job_code"], (None, None, None))
        p["city"]  = c or p.pop("erp_city",  None)
        p["state"] = s or p.pop("erp_state", None)
        p["zip"]   = z or p.pop("erp_zip",   None)

    _PN_KEYS = ["part_type", "subcategory", "generation", "diameter", "height",
               "opening_diameter", "troughing", "wall_variant", "section_suffix",
               "lid_suffix", "box_length", "box_suffix", "es", "de", "de_count"]
    PN_ATTRS = [f"pn_{k}" for k in _PN_KEYS]
    fields = ["piece_id", "job_code", "structure_id", "part_number", "gen4_name", "part_type",
              "structure_class", "ship_date", "year", "plant", "source", "city", "state", "zip",
              *PN_ATTRS]
    path = os.path.join(sources.OUTPUT_DIR, "pieces_ledger.csv")
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for p in ledger:
            row = {k: p.get(k) for k in fields}
            attrs = parse_part_name(p.get("part_number"))
            for col, key in zip(PN_ATTRS, _PN_KEYS):
                row[col] = attrs.get(key, '')
            row["gen4_name"] = build_gen4_name(attrs)
            w.writerow(row)

    assert len({p["piece_id"] for p in ledger}) == len(ledger), "duplicate piece_id!"
    cls = Counter(p["structure_class"] for p in ledger)
    print(f"ledger -> {path}")
    print(f"class: {dict(cls)} | sum={sum(cls.values())}")

    shipped_jobs = {p["job_code"] for p in ledger if p["job_code"]}
    qpath, qn = quoted_not_shipped(shipped_jobs)
    print(f"quoted_not_shipped: {qn} jobs -> {qpath}")
    return ledger


if __name__ == "__main__":
    build()
