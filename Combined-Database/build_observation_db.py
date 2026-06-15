import csv
import json
import os
import sqlite3
from collections import defaultdict

import openpyxl

import sources
import trust_model
from common import (norm_city, norm_county, norm_state, norm_zip, norm_job_code,
                    year_from_date, is_blank, is_county_value)

csv.field_size_limit(10_000_000)

# Per-source job-level fields collected as: agg[(jc, field, source)][value_norm] = (count, value_raw)
LOC_FIELDS = ("city", "state", "zip", "county", "project_name", "customer", "plant", "year")


def _add(agg, jc, field, source, value_norm, value_raw):
    if not jc or value_norm is None:
        return
    key = (jc, field, source)
    cur = agg[key].get(value_norm)
    agg[key][value_norm] = (cur[0] + 1 if cur else 1, value_raw if value_raw is not None else value_norm)


def read_bom_union(agg):
    wb = openpyxl.load_workbook(sources.BOM_UNION_XLSX, read_only=True, data_only=True)
    ws = wb[sources.BOM_UNION_SHEET]
    it = ws.iter_rows(values_only=True)
    h = {str(c).strip(): i for i, c in enumerate(next(it)) if c is not None}
    for row in it:
        jc = norm_job_code(row[h["Job Code"]])
        if not jc:
            continue
        loc = row[h["Job Location"]]
        if isinstance(loc, str) and "," in loc:
            city_raw, state_raw = loc.rsplit(",", 1)
            state = norm_state(state_raw)
            if state:
                if is_county_value(city_raw):
                    _add(agg, jc, "county", "bom_union", norm_county(city_raw), loc)
                else:
                    _add(agg, jc, "city", "bom_union", norm_city(city_raw), loc)
                _add(agg, jc, "state", "bom_union", state, loc)
            else:
                # state parse failed — full string is a street address, keep for geocoding
                _add(agg, jc, "street_address", "bom_union", loc.strip(), loc)
        _add(agg, jc, "zip", "bom_union", norm_zip(row[h["Zip Code"]]), row[h["Zip Code"]])
        _add(agg, jc, "customer", "bom_union",
             (str(row[h["Contractor"]]).strip().upper() if not is_blank(row[h["Contractor"]]) else None),
             row[h["Contractor"]])
        _add(agg, jc, "year", "bom_union", year_from_date(row[h["Year Release"]]), row[h["Year Release"]])
        if "Project Name" in h:
            pn = row[h["Project Name"]]
            if not is_blank(pn):
                _add(agg, jc, "project_name", "bom_union", str(pn).strip().upper(), pn)
    wb.close()


def read_dispatch(agg):
    with open(sources.DISPATCH_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            jc = norm_job_code(row.get("job_code"))
            if not jc:
                continue
            raw_city = row.get("city")
            if is_county_value(raw_city):
                _add(agg, jc, "county", "dispatch", norm_county(raw_city), raw_city)
            else:
                _add(agg, jc, "city", "dispatch", norm_city(raw_city), raw_city)
            _add(agg, jc, "state", "dispatch", norm_state(row.get("state")), row.get("state"))
            plant = (row.get("plant") or "").strip().upper() or None
            _add(agg, jc, "plant", "dispatch", plant, row.get("plant"))
            _add(agg, jc, "year", "dispatch", year_from_date(row.get("year") or row.get("ship_date")), row.get("year"))
            job_name = row.get("job_name")
            if not is_blank(job_name):
                _add(agg, jc, "project_name", "dispatch", str(job_name).strip().upper(), job_name)


def read_baby(agg):
    wb = openpyxl.load_workbook(sources.BABY_XLSM, read_only=True, data_only=True)
    ws = wb[sources.BABY_SHEET]
    it = ws.iter_rows(values_only=True)
    h = {str(c).strip(): i for i, c in enumerate(next(it)) if c is not None}
    jc_i = h["Job Code"]
    for row in it:
        jc = norm_job_code(row[jc_i])
        if not jc:
            continue
        src = sources.ERP_LABELS.get(str(row[h["ERP"]]).strip()) if row[h["ERP"]] else None
        if not src:
            continue
        _add(agg, jc, "city", src, norm_city(row[h["Shipping City"]]), row[h["Shipping City"]])
        _add(agg, jc, "state", src, norm_state(row[h["Shippings State"]]), row[h["Shippings State"]])
        _add(agg, jc, "zip", src, norm_zip(row[h["ZipCode"]]), row[h["ZipCode"]])
        cust = row[h["Invoiced Custumer"]]
        _add(agg, jc, "customer", src,
             (str(cust).strip().upper() if not is_blank(cust) else None), cust)
        plant = (str(row[h["Plant"]]).strip().upper() if not is_blank(row[h["Plant"]]) else None)
        _add(agg, jc, "plant", src, plant, row[h["Plant"]])
        _add(agg, jc, "year", src, year_from_date(row[h["Date Shipped"]]), row[h["Date Shipped"]])
        if "Shipping County" in h:
            _add(agg, jc, "county", src, norm_county(row[h["Shipping County"]]), row[h["Shipping County"]])
    wb.close()


def read_jobcode_db(agg):
    recs = json.load(open(sources.JOBCODE_DB_JSON, encoding="utf-8"))
    for r in recs:
        jc = norm_job_code(r.get("job_code"))
        if not jc:
            continue
        _add(agg, jc, "county", "jobcode_db", norm_county(r.get("shipping_county")), r.get("shipping_county"))
        raw_city = r.get("shipping_city")
        if is_county_value(raw_city):
            _add(agg, jc, "county", "jobcode_db", norm_county(raw_city), raw_city)
        else:
            _add(agg, jc, "city", "jobcode_db", norm_city(raw_city), raw_city)
        _add(agg, jc, "state", "jobcode_db", norm_state(r.get("shipping_state")), r.get("shipping_state"))
        _add(agg, jc, "zip", "jobcode_db", norm_zip(r.get("shipping_zip")), r.get("shipping_zip"))
        _add(agg, jc, "customer", "jobcode_db",
             (str(r.get("customer")).strip().upper() if not is_blank(r.get("customer")) else None), r.get("customer"))
        _add(agg, jc, "plant", "jobcode_db",
             (str(r.get("plant")).strip().upper() if not is_blank(r.get("plant")) else None), r.get("plant"))
        _add(agg, jc, "year", "jobcode_db", year_from_date(r.get("year_released")), r.get("year_released"))
        pn = r.get("project_name")
        if not is_blank(pn):
            _add(agg, jc, "project_name", "jobcode_db", str(pn).strip().upper(), pn)


def read_registry(agg):
    wb = openpyxl.load_workbook(sources.JOB_CODE_REGISTRY_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    h = {str(c).strip(): i for i, c in enumerate(next(it)) if c is not None}
    for row in it:
        jc = norm_job_code(row[h["Job Code"]])
        if not jc:
            continue
        if "County" in h:
            _add(agg, jc, "county", "registry", norm_county(row[h["County"]]), row[h["County"]])
        raw_city = row[h["City"]]
        if is_county_value(raw_city):
            _add(agg, jc, "county", "registry", norm_county(raw_city), raw_city)
        else:
            _add(agg, jc, "city", "registry", norm_city(raw_city), raw_city)
        _add(agg, jc, "state", "registry", norm_state(row[h["State"]]), row[h["State"]])
        cust = row[h["Shipping Customer"]]
        _add(agg, jc, "customer", "registry",
             (str(cust).strip().upper() if not is_blank(cust) else None), cust)
        if "BOM Project Name" in h:
            pn = row[h["BOM Project Name"]]
            if not is_blank(pn):
                _add(agg, jc, "project_name", "registry", str(pn).strip().upper(), pn)
    wb.close()


def build():
    sources.ensure_dirs()
    agg = defaultdict(dict)
    print("reading bom_union...");   read_bom_union(agg)
    print("reading dispatch...");    read_dispatch(agg)
    print("reading baby (erp)...");  read_baby(agg)
    print("reading jobcode_db...");  read_jobcode_db(agg)
    print("reading registry...");    read_registry(agg)

    if os.path.exists(sources.OBSERVATIONS_DB):
        os.remove(sources.OBSERVATIONS_DB)
    con = sqlite3.connect(sources.OBSERVATIONS_DB)
    con.execute("""CREATE TABLE observations(
        obs_id TEXT PRIMARY KEY, entity_type TEXT, entity_key TEXT, field TEXT,
        value_raw TEXT, value_norm TEXT, source TEXT, source_detail TEXT,
        trust REAL, n_pieces INTEGER)""")

    rows = []
    seq = defaultdict(int)
    for (jc, field, source), values in agg.items():
        for value_norm, (count, value_raw) in values.items():
            seq_key = (jc, field, source)
            seq[seq_key] += 1
            obs_id = f"{jc}::{field}::{source}::{seq[seq_key]}"
            rows.append((obs_id, "job_code", jc, field,
                         str(value_raw) if value_raw is not None else None, str(value_norm),
                         source, source, trust_model.score(field, source), count))
    con.executemany("INSERT INTO observations VALUES (?,?,?,?,?,?,?,?,?,?)", rows)
    con.execute("CREATE INDEX ix_obs_ef ON observations(entity_key, field)")

    # entity_field_resolution: one row per (entity_key, field) with status + winner.
    con.execute("""CREATE TABLE entity_field_resolution AS
        WITH per_value AS (
            SELECT entity_key, field, value_norm,
                   SUM(trust) AS trust_sum, SUM(n_pieces) AS pieces,
                   GROUP_CONCAT(DISTINCT source) AS srcs
            FROM observations GROUP BY entity_key, field, value_norm),
        ranked AS (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY entity_key, field
                     ORDER BY trust_sum DESC, pieces DESC) AS rnk
            FROM per_value),
        counts AS (
            SELECT entity_key, field, COUNT(*) AS n_values FROM per_value
            GROUP BY entity_key, field)
        SELECT r.entity_key, r.field,
               CASE WHEN c.n_values > 1 THEN 'CONFLICT' ELSE 'AGREE' END AS status,
               r.value_norm AS winner, r.srcs AS winner_sources, r.trust_sum AS winner_trust,
               c.n_values
        FROM ranked r JOIN counts c
          ON r.entity_key = c.entity_key AND r.field = c.field
        WHERE r.rnk = 1""")
    con.commit()

    n_obs = con.execute("SELECT COUNT(*) FROM observations").fetchone()[0]
    n_ent = con.execute("SELECT COUNT(DISTINCT entity_key) FROM observations").fetchone()[0]
    n_conf = con.execute("SELECT COUNT(*) FROM entity_field_resolution WHERE status='CONFLICT'").fetchone()[0]
    con.close()
    print(f"observations: {n_obs} | job_codes: {n_ent} | conflicts: {n_conf}")
    print("wrote", sources.OBSERVATIONS_DB)


if __name__ == "__main__":
    build()
